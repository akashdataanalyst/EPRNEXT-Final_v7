import csv
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


COMPANY_NAME = "Calco PolyTechnik Pvt Ltd"
STOCK_UOM = "Kg"
FG_ITEM_GROUP_ROOT = "Finished Good"
RM_ITEM_GROUP_ROOT = "Raw Material"
BOM_TOTAL_TOLERANCE = 0.01

RM_CATEGORY_MAP = {
    "Polymer PA": "Polymer PA",
    "Polymer PBT": "Polymer PBT",
    "Polymer PE/PP": "Polymer PE PP",
    "Polymer PC": "Polymer PC",
    "Polymer Styrene": "Polymer Styrene",
    "Mineral": "Mineral",
    "Glass Fibre": "Glass Fibre",
    "Additive": "Additive",
    "Pigment": "Pigment",
    "Packing": "Packing Material",
}

FG_CATEGORY_MAP = {
    "PA6": "PA6",
    "PA66": "PA66",
    "PBT": "PBT",
    "PP": "PP",
    "ABS": "ABS",
    "Alloy": "Alloy",
}

QUALITY_TEMPLATE_MAP = {
    "Finished Good": "Calco Final FG QC",
    "Raw Material": "Calco Incoming RM QC",
}

WAREHOUSE_ROWS = [
    {"warehouse_name": "Raw Material - HOLD", "parent_warehouse": "", "company": COMPANY_NAME},
    {"warehouse_name": "Raw Material - RELEASED", "parent_warehouse": "", "company": COMPANY_NAME},
    {"warehouse_name": "Raw Material - REJECTED", "parent_warehouse": "", "company": COMPANY_NAME},
    {"warehouse_name": "WIP - Production", "parent_warehouse": "", "company": COMPANY_NAME},
    {"warehouse_name": "Finished Good - HOLD", "parent_warehouse": "", "company": COMPANY_NAME},
    {"warehouse_name": "Finished Good - RELEASED", "parent_warehouse": "", "company": COMPANY_NAME},
    {"warehouse_name": "Finished Good - REJECTED", "parent_warehouse": "", "company": COMPANY_NAME},
]

ITEM_GROUP_ROWS = [
    {"item_group_name": RM_ITEM_GROUP_ROOT, "parent_item_group": "All Item Groups", "is_group": 1},
    {"item_group_name": "Polymer", "parent_item_group": RM_ITEM_GROUP_ROOT, "is_group": 1},
    {"item_group_name": "Polymer PA", "parent_item_group": "Polymer", "is_group": 0},
    {"item_group_name": "Polymer PBT", "parent_item_group": "Polymer", "is_group": 0},
    {"item_group_name": "Polymer PE PP", "parent_item_group": "Polymer", "is_group": 0},
    {"item_group_name": "Polymer PC", "parent_item_group": "Polymer", "is_group": 0},
    {"item_group_name": "Polymer Styrene", "parent_item_group": "Polymer", "is_group": 0},
    {"item_group_name": "Additive", "parent_item_group": RM_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "Pigment", "parent_item_group": RM_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "Mineral", "parent_item_group": RM_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "Glass Fibre", "parent_item_group": RM_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "Packing Material", "parent_item_group": RM_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "Unclassified RM", "parent_item_group": RM_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": FG_ITEM_GROUP_ROOT, "parent_item_group": "All Item Groups", "is_group": 1},
    {"item_group_name": "PA6", "parent_item_group": FG_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "PA66", "parent_item_group": FG_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "PBT", "parent_item_group": FG_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "PP", "parent_item_group": FG_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "ABS", "parent_item_group": FG_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "Alloy", "parent_item_group": FG_ITEM_GROUP_ROOT, "is_group": 0},
    {"item_group_name": "Unclassified FG", "parent_item_group": FG_ITEM_GROUP_ROOT, "is_group": 0},
]


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_code(value):
    return normalize_text(value).upper()


def clean_item_name(value, fallback_code):
    name = normalize_text(value)
    return name or fallback_code


def safe_float(value):
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number):
        return None
    return number


def parse_sheet(ws):
    headers = [normalize_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        payload = {headers[idx]: row[idx] for idx in range(len(headers))}
        rows.append(payload)
    return rows


def row_key(row, fieldname):
    return normalize_code(row.get(fieldname))


def classify_rm_group(category):
    category = normalize_text(category)
    return RM_CATEGORY_MAP.get(category, "Unclassified RM")


def classify_fg_group(category):
    category = normalize_text(category)
    return FG_CATEGORY_MAP.get(category, "Unclassified FG")


def build_item_row(row, item_type):
    code = normalize_code(row.get("SKU Code"))
    category = normalize_text(row.get("Category"))
    item_name = clean_item_name(row.get("Item Name"), code)
    item_group = classify_rm_group(category) if item_type == "RM" else classify_fg_group(category)
    description = normalize_text(row.get("Item Details\n(Optional)"))
    return {
        "item_code": code,
        "item_name": item_name,
        "item_group": item_group,
        "stock_uom": STOCK_UOM,
        "is_stock_item": 1,
        "include_item_in_manufacturing": 1,
        "valuation_method": "Moving Average" if item_type == "RM" else "FIFO",
        "has_batch_no": 1,
        "create_new_batch": 1 if item_type == "FG" else 0,
        "inspection_required_before_purchase": 1 if item_type == "RM" else 0,
        "inspection_required_before_delivery": 1 if item_type == "FG" else 0,
        "custom_enable_rm_qc": 1 if item_type == "RM" else 0,
        "allow_alternative_item": 0,
        "disabled": 0,
        "description": description,
        "quality_template": QUALITY_TEMPLATE_MAP["Raw Material" if item_type == "RM" else "Finished Good"],
    }


def build_uom_rows(rm_rows, fg_rows):
    rows = []
    for item_type, source_rows in (("RM", rm_rows), ("FG", fg_rows)):
        for row in source_rows:
            code = normalize_code(row.get("SKU Code"))
            if not code:
                continue
            rows.append(
                {
                    "item_code": code,
                    "item_type": item_type,
                    "source_category": normalize_text(row.get("Category")),
                    "stock_uom": STOCK_UOM,
                    "purchase_uom": STOCK_UOM,
                    "conversion_factor": 1,
                }
            )
    return rows


def build_duplicate_report(rows, fieldname, sheet_name):
    counts = Counter(row_key(row, fieldname) for row in rows if row_key(row, fieldname))
    report = []
    for code, count in sorted(counts.items()):
        if count > 1:
            report.append({"source_sheet": sheet_name, "code": code, "occurrences": count})
    return report


def build_formulation_groups(formulation_rows):
    grouped = defaultdict(list)
    for row in formulation_rows:
        product_code = normalize_code(row.get("Product Code"))
        revision = normalize_text(row.get("Revision Code"))
        grouped[(product_code, revision)].append(row)
    return grouped


def build_exception_reports(rm_rows, fg_rows, formulation_rows):
    rm_codes = {row_key(row, "SKU Code") for row in rm_rows if row_key(row, "SKU Code")}
    fg_codes = {row_key(row, "SKU Code") for row in fg_rows if row_key(row, "SKU Code")}

    duplicate_fg = build_duplicate_report(fg_rows, "SKU Code", "FG inventory")
    duplicate_rm = build_duplicate_report(rm_rows, "SKU Code", "RM inventory")

    missing_fg = []
    missing_rm = []
    invalid_formulations = []
    bom_mismatches = []
    exception_log = []

    grouped = build_formulation_groups(formulation_rows)

    for row in formulation_rows:
        product_code = normalize_code(row.get("Product Code"))
        revision = normalize_text(row.get("Revision Code"))
        rm_code = normalize_code(row.get("RM Code"))
        dosage = safe_float(row.get("Dosage (%)"))

        if product_code and product_code not in fg_codes:
            missing_fg.append(
                {"product_code": product_code, "revision_code": revision, "issue": "FG code missing in FG inventory"}
            )

        if not rm_code:
            missing_rm.append(
                {"product_code": product_code, "revision_code": revision, "rm_code": "", "issue": "RM code blank"}
            )
        elif rm_code not in rm_codes:
            missing_rm.append(
                {
                    "product_code": product_code,
                    "revision_code": revision,
                    "rm_code": rm_code,
                    "issue": "RM code missing in RM inventory",
                }
            )

        if dosage is None or dosage <= 0:
            invalid_formulations.append(
                {
                    "product_code": product_code,
                    "revision_code": revision,
                    "rm_code": rm_code,
                    "dosage_percent": row.get("Dosage (%)"),
                    "issue": "Dosage must be numeric and greater than zero",
                }
            )

    for (product_code, revision), rows in sorted(grouped.items()):
        total = 0.0
        valid_line_count = 0
        for row in rows:
            dosage = safe_float(row.get("Dosage (%)"))
            if dosage and dosage > 0:
                total += dosage
                valid_line_count += 1

        if not product_code:
            continue

        if valid_line_count == 0 or abs(total - 100.0) > BOM_TOTAL_TOLERANCE:
            bom_mismatches.append(
                {
                    "product_code": product_code,
                    "revision_code": revision,
                    "valid_line_count": valid_line_count,
                    "total_dosage_percent": round(total, 6),
                }
            )

    for row in duplicate_fg:
        exception_log.append(
            {
                "severity": "Error",
                "source": row["source_sheet"],
                "record_code": row["code"],
                "issue_type": "Duplicate FG Code",
                "details": f"FG code appears {row['occurrences']} times",
            }
        )
    for row in duplicate_rm:
        exception_log.append(
            {
                "severity": "Error",
                "source": row["source_sheet"],
                "record_code": row["code"],
                "issue_type": "Duplicate RM Code",
                "details": f"RM code appears {row['occurrences']} times",
            }
        )
    for row in missing_fg:
        exception_log.append(
            {
                "severity": "Error",
                "source": "Formulations",
                "record_code": row["product_code"],
                "issue_type": "Missing FG Reference",
                "details": row["issue"],
            }
        )
    for row in missing_rm:
        exception_log.append(
            {
                "severity": "Error",
                "source": "Formulations",
                "record_code": row["rm_code"] or f"{row['product_code']}:{row['revision_code']}",
                "issue_type": "Missing RM Reference",
                "details": row["issue"],
            }
        )
    for row in invalid_formulations:
        exception_log.append(
            {
                "severity": "Error",
                "source": "Formulations",
                "record_code": f"{row['product_code']}:{row['revision_code']}",
                "issue_type": "Invalid Formulation Value",
                "details": row["issue"],
            }
        )
    for row in bom_mismatches:
        exception_log.append(
            {
                "severity": "Error",
                "source": "Formulations",
                "record_code": f"{row['product_code']}:{row['revision_code']}",
                "issue_type": "BOM Total Mismatch",
                "details": f"Valid dosage total = {row['total_dosage_percent']}",
            }
        )

    return {
        "duplicate_fg_codes": duplicate_fg,
        "duplicate_rm_codes": duplicate_rm,
        "missing_fg_references": missing_fg,
        "missing_rm_references": missing_rm,
        "invalid_formulation_values": invalid_formulations,
        "bom_total_mismatches": bom_mismatches,
        "exception_log": exception_log,
    }


def build_bom_templates(fg_rows, formulation_rows, exceptions):
    fg_lookup = {}
    for row in fg_rows:
        code = row_key(row, "SKU Code")
        if code and code not in fg_lookup:
            fg_lookup[code] = row

    missing_fg_keys = {
        (row["product_code"], row["revision_code"]) for row in exceptions["missing_fg_references"]
    }
    missing_rm_keys = {
        (row["product_code"], row["revision_code"]) for row in exceptions["missing_rm_references"]
    }
    invalid_keys = {
        (row["product_code"], row["revision_code"]) for row in exceptions["invalid_formulation_values"]
    }
    mismatch_keys = {
        (row["product_code"], row["revision_code"]) for row in exceptions["bom_total_mismatches"]
    }

    grouped = build_formulation_groups(formulation_rows)
    revisions_by_fg = defaultdict(list)
    for product_code, revision in grouped:
        revisions_by_fg[product_code].append(revision)

    bom_headers = []
    bom_items = []

    for product_code, revisions in sorted(revisions_by_fg.items()):
        active_revision = max(
            revisions,
            key=lambda value: (safe_float(value) if safe_float(value) is not None else -1, str(value)),
        )

        for revision in sorted(revisions, key=lambda value: str(value)):
            key = (product_code, revision)
            if key in missing_fg_keys or key in missing_rm_keys or key in invalid_keys or key in mismatch_keys:
                continue

            fg_row = fg_lookup.get(product_code)
            if not fg_row:
                continue

            bom_name = f"BOM-{product_code}-R{revision}"
            bom_headers.append(
                {
                    "name": bom_name,
                    "item": product_code,
                    "item_name": clean_item_name(fg_row.get("Item Name"), product_code),
                    "company": COMPANY_NAME,
                    "quantity": 100,
                    "uom": STOCK_UOM,
                    "is_active": 1 if revision == active_revision else 0,
                    "is_default": 1 if revision == active_revision else 0,
                    "allow_alternative_item": 0,
                    "set_rate_of_sub_assembly_item_based_on_bom": 0,
                    "rm_cost_as_per": "Valuation Rate",
                    "with_operations": 0,
                    "fg_based_operating_cost": 0,
                    "project": "",
                }
            )

            for row in grouped[key]:
                rm_code = normalize_code(row.get("RM Code"))
                dosage = safe_float(row.get("Dosage (%)"))
                if not rm_code or dosage is None or dosage <= 0:
                    continue
                bom_items.append(
                    {
                        "parent": bom_name,
                        "parenttype": "BOM",
                        "parentfield": "items",
                        "item_code": rm_code,
                        "qty": round(dosage, 6),
                        "uom": STOCK_UOM,
                        "stock_qty": round(dosage, 6),
                        "stock_uom": STOCK_UOM,
                        "include_item_in_manufacturing": 1,
                    }
                )

    return bom_headers, bom_items


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def build_outputs(workbook_path, output_dir):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rm_rows = parse_sheet(workbook["RM inventory"])
    fg_rows = parse_sheet(workbook["FG inventory"])
    formulation_rows = parse_sheet(workbook["Formulations"])

    exceptions = build_exception_reports(rm_rows, fg_rows, formulation_rows)
    bom_headers, bom_items = build_bom_templates(fg_rows, formulation_rows, exceptions)

    unique_rm_rows = []
    seen_rm = set()
    for row in rm_rows:
        code = row_key(row, "SKU Code")
        if not code or code in seen_rm:
            continue
        unique_rm_rows.append(build_item_row(row, "RM"))
        seen_rm.add(code)

    unique_fg_rows = []
    seen_fg = set()
    for row in fg_rows:
        code = row_key(row, "SKU Code")
        if not code or code in seen_fg:
            continue
        unique_fg_rows.append(build_item_row(row, "FG"))
        seen_fg.add(code)

    outputs = {
        "item_groups.csv": (ITEM_GROUP_ROWS, ["item_group_name", "parent_item_group", "is_group"]),
        "uom_normalization.csv": (
            build_uom_rows(rm_rows, fg_rows),
            ["item_code", "item_type", "source_category", "stock_uom", "purchase_uom", "conversion_factor"],
        ),
        "warehouses.csv": (
            WAREHOUSE_ROWS,
            ["warehouse_name", "parent_warehouse", "company"],
        ),
        "items_rm.csv": (
            unique_rm_rows,
            [
                "item_code",
                "item_name",
                "item_group",
                "stock_uom",
                "is_stock_item",
                "include_item_in_manufacturing",
                "valuation_method",
                "has_batch_no",
                "create_new_batch",
                "inspection_required_before_purchase",
                "inspection_required_before_delivery",
                "custom_enable_rm_qc",
                "allow_alternative_item",
                "disabled",
                "description",
                "quality_template",
            ],
        ),
        "items_fg.csv": (
            unique_fg_rows,
            [
                "item_code",
                "item_name",
                "item_group",
                "stock_uom",
                "is_stock_item",
                "include_item_in_manufacturing",
                "valuation_method",
                "has_batch_no",
                "create_new_batch",
                "inspection_required_before_purchase",
                "inspection_required_before_delivery",
                "custom_enable_rm_qc",
                "allow_alternative_item",
                "disabled",
                "description",
                "quality_template",
            ],
        ),
        "bom_header.csv": (
            bom_headers,
            [
                "name",
                "item",
                "item_name",
                "company",
                "quantity",
                "uom",
                "is_active",
                "is_default",
                "allow_alternative_item",
                "set_rate_of_sub_assembly_item_based_on_bom",
                "rm_cost_as_per",
                "with_operations",
                "fg_based_operating_cost",
                "project",
            ],
        ),
        "bom_items.csv": (
            bom_items,
            [
                "parent",
                "parenttype",
                "parentfield",
                "item_code",
                "qty",
                "uom",
                "stock_qty",
                "stock_uom",
                "include_item_in_manufacturing",
            ],
        ),
        "duplicate_fg_codes.csv": (
            exceptions["duplicate_fg_codes"],
            ["source_sheet", "code", "occurrences"],
        ),
        "duplicate_rm_codes.csv": (
            exceptions["duplicate_rm_codes"],
            ["source_sheet", "code", "occurrences"],
        ),
        "missing_fg_references.csv": (
            exceptions["missing_fg_references"],
            ["product_code", "revision_code", "issue"],
        ),
        "missing_rm_references.csv": (
            exceptions["missing_rm_references"],
            ["product_code", "revision_code", "rm_code", "issue"],
        ),
        "invalid_formulation_values.csv": (
            exceptions["invalid_formulation_values"],
            ["product_code", "revision_code", "rm_code", "dosage_percent", "issue"],
        ),
        "bom_total_mismatches.csv": (
            exceptions["bom_total_mismatches"],
            ["product_code", "revision_code", "valid_line_count", "total_dosage_percent"],
        ),
        "exception_log.csv": (
            exceptions["exception_log"],
            ["severity", "source", "record_code", "issue_type", "details"],
        ),
    }

    output_dir = Path(output_dir)
    for filename, (rows, fieldnames) in outputs.items():
        write_csv(output_dir / filename, rows, fieldnames)

    summary = {
        "rm_inventory_rows": len(rm_rows),
        "fg_inventory_rows": len(fg_rows),
        "formulation_rows": len(formulation_rows),
        "unique_rm_codes": len(unique_rm_rows),
        "unique_fg_codes": len(unique_fg_rows),
        "valid_bom_headers": len(bom_headers),
        "valid_bom_items": len(bom_items),
        "duplicate_fg_codes": len(exceptions["duplicate_fg_codes"]),
        "duplicate_rm_codes": len(exceptions["duplicate_rm_codes"]),
        "missing_fg_references": len(exceptions["missing_fg_references"]),
        "missing_rm_references": len(exceptions["missing_rm_references"]),
        "invalid_formulation_values": len(exceptions["invalid_formulation_values"]),
        "bom_total_mismatches": len(exceptions["bom_total_mismatches"]),
    }
    write_json(output_dir / "summary.json", summary)
    return summary


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Generate Calco ERP master-data import templates.")
    parser.add_argument("workbook_path", help="Path to the source Excel workbook")
    parser.add_argument("output_dir", help="Directory where generated CSVs should be written")
    args = parser.parse_args()

    summary = build_outputs(args.workbook_path, args.output_dir)
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
