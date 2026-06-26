from __future__ import annotations

from pathlib import Path

import frappe
from frappe import _
from openpyxl import load_workbook

from calco_erp.data_foundation.import_master_data import ensure_master_prerequisites
from calco_erp.data_foundation.master_data_builder import clean_item_name, normalize_code, normalize_text


RAW_MATERIAL_GROUP = "Raw Material"
RM_QUALITY_TEMPLATE = "Calco Incoming RM QC"
LEGACY_ALIAS_MAP = {
    # Existing live ERP code already represents this source RM.
    "L11XM": "2110",
}
DEFAULT_WORKBOOK_PATH = "/tmp/rm_ims.xlsx"


def _parse_rm_rows(workbook_path: str) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_text(cell) for cell in header_cells]

    rows = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: values[index] for index in range(len(headers))}
        item_code = normalize_code(row.get("SKU Code"))
        if not item_code:
            continue
        rows.append(
            {
                "item_code": item_code,
                "category": normalize_text(row.get("Category")),
                "item_name": clean_item_name(row.get("Item Name"), item_code),
                "description": normalize_text(row.get("Item Details (Optional)") or row.get("Item Details\n(Optional)")),
            }
        )
    return rows


def _build_source_index(workbook_path: str) -> dict[str, dict[str, str]]:
    return {row["item_code"]: row for row in _parse_rm_rows(workbook_path)}


def _build_live_item_index() -> dict[str, dict[str, object]]:
    rows = frappe.get_all(
        "Item",
        fields=["name", "item_name", "item_group", "disabled", "is_stock_item", "description"],
        limit_page_length=0,
    )
    return {normalize_code(row["name"]): row for row in rows}


def _build_missing_report(source_index: dict[str, dict[str, str]], live_index: dict[str, dict[str, object]]) -> list[dict[str, object]]:
    report = []
    for item_code, source_row in sorted(source_index.items()):
        live_row = live_index.get(item_code)
        alias_code = LEGACY_ALIAS_MAP.get(item_code)
        alias_row = live_index.get(alias_code) if alias_code else None

        if live_row and normalize_text(live_row.get("item_group")) == RAW_MATERIAL_GROUP:
            continue

        exists_in_erp = bool(live_row or alias_row)
        current_group = normalize_text((live_row or alias_row or {}).get("item_group"))
        disabled = (live_row or alias_row or {}).get("disabled")
        is_stock_item = (live_row or alias_row or {}).get("is_stock_item")

        if live_row and normalize_text(live_row.get("item_group")) != RAW_MATERIAL_GROUP:
            reason = "Wrong item group"
        elif alias_row:
            reason = "Duplicate/merged code"
        else:
            reason = "Not imported"

        report.append(
            {
                "item_code": item_code,
                "item_name": source_row["item_name"],
                "exists_in_erp": "Yes" if exists_in_erp else "No",
                "current_item_group": current_group,
                "disabled": disabled,
                "is_stock_item": is_stock_item,
                "missing_reason": reason,
                "matched_erp_code": alias_code or "",
            }
        )
    return report


def _ensure_rm_item(source_row: dict[str, str]) -> str:
    item_code = source_row["item_code"]
    item = frappe.new_doc("Item")
    item.item_code = item_code
    item.item_name = source_row["item_name"] or item_code
    item.item_group = RAW_MATERIAL_GROUP
    item.stock_uom = "Kg"
    item.is_stock_item = 1
    item.include_item_in_manufacturing = 1
    item.valuation_method = "Moving Average"
    item.has_batch_no = 1
    item.inspection_required_before_purchase = 1
    item.quality_inspection_template = RM_QUALITY_TEMPLATE
    item.custom_enable_rm_qc = 1
    item.allow_alternative_item = 0
    item.disabled = 0
    item.description = source_row.get("description") or ""
    item.insert(ignore_permissions=True)
    return item.name


def _reclassify_rm_item(item_code: str, source_row: dict[str, str]) -> str:
    item = frappe.get_doc("Item", item_code)
    item.item_group = RAW_MATERIAL_GROUP
    item.item_name = source_row["item_name"] or item.item_name or item_code
    item.stock_uom = item.stock_uom or "Kg"
    item.is_stock_item = 1
    item.include_item_in_manufacturing = 1
    item.valuation_method = item.valuation_method or "Moving Average"
    item.has_batch_no = 1
    item.inspection_required_before_purchase = 1
    item.quality_inspection_template = item.quality_inspection_template or RM_QUALITY_TEMPLATE
    item.custom_enable_rm_qc = 1
    item.disabled = 0
    if source_row.get("description"):
        item.description = source_row["description"]
    item.save(ignore_permissions=True)
    return item.name


def _resolve_workbook_path(workbook_path: str | None) -> str:
    return str(Path(workbook_path or DEFAULT_WORKBOOK_PATH))


@frappe.whitelist()
def generate_cleanup_report(workbook_path: str | None = None) -> dict[str, object]:
    ensure_master_prerequisites()
    workbook_path = _resolve_workbook_path(workbook_path)
    source_index = _build_source_index(workbook_path)
    live_index = _build_live_item_index()
    missing_report = _build_missing_report(source_index, live_index)

    raw_material_count = frappe.db.count("Item", {"item_group": RAW_MATERIAL_GROUP, "disabled": 0})
    erp_only = []
    for item_code, live_row in sorted(live_index.items()):
        if normalize_text(live_row.get("item_group")) != RAW_MATERIAL_GROUP:
            continue
        if item_code not in source_index and item_code not in LEGACY_ALIAS_MAP.values():
            erp_only.append(
                {
                    "item_code": live_row["name"],
                    "item_name": live_row["item_name"],
                    "disabled": live_row["disabled"],
                    "is_stock_item": live_row["is_stock_item"],
                }
            )

    return {
        "source_rm_count": len(source_index),
        "erp_rm_count": raw_material_count,
        "missing_report": missing_report,
        "erp_only_rm_items": erp_only,
        "legacy_alias_map": LEGACY_ALIAS_MAP,
        "notes": [
            "RM IMS.xlsx is treated as a migration reference only.",
            "No ERP items are automatically disabled or end-of-lifed by this cleanup.",
            "Legacy alias codes are reported for review and left active unless explicitly changed later.",
        ],
    }


@frappe.whitelist()
def apply_cleanup(workbook_path: str | None = None) -> dict[str, object]:
    ensure_master_prerequisites()
    workbook_path = _resolve_workbook_path(workbook_path)
    source_index = _build_source_index(workbook_path)
    live_index = _build_live_item_index()
    missing_report = _build_missing_report(source_index, live_index)

    created_items = []
    reclassified_items = []
    reviewed_aliases = []

    for row in missing_report:
        source_row = source_index[row["item_code"]]
        if row["missing_reason"] == "Wrong item group":
            _reclassify_rm_item(row["item_code"], source_row)
            reclassified_items.append(row["item_code"])
        elif row["missing_reason"] == "Duplicate/merged code":
            if row["matched_erp_code"] and row["matched_erp_code"] in LEGACY_ALIAS_MAP.values():
                reviewed_aliases.append(
                    {
                        "source_item_code": row["item_code"],
                        "erp_item_code": row["matched_erp_code"],
                        "action": "retained existing ERP alias",
                    }
                )
            else:
                _ensure_rm_item(source_row)
                created_items.append(row["item_code"])
        else:
            _ensure_rm_item(source_row)
            created_items.append(row["item_code"])

    frappe.db.commit()

    final_report = generate_cleanup_report(workbook_path)
    final_report.update(
        {
            "created_items": created_items,
            "reclassified_items": reclassified_items,
            "reviewed_aliases": reviewed_aliases,
            "final_erp_rm_count": frappe.db.count("Item", {"item_group": RAW_MATERIAL_GROUP, "disabled": 0}),
        }
    )
    return final_report
