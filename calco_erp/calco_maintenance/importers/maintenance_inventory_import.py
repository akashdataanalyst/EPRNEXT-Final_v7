from __future__ import annotations

import json
from pathlib import Path

import frappe
from openpyxl import load_workbook

from calco_erp.calco_maintenance.master_data_utils import cstr, get_frequency_details, guess_equipment_group


DEFAULT_SPARE_UOM = "Nos"
MAINTENANCE_SPARE_GROUP = "Maintenance Spares"


def generated_dir():
    return Path("/tmp/maintenance_inventory_import")


def iter_non_empty_rows(sheet):
    for row in sheet.iter_rows(values_only=True):
        values = [cstr(value) for value in row]
        if any(values):
            yield values


def read_workbook(workbook_path: str | Path) -> dict[str, list[dict[str, str]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    equipment_rows = []
    spare_rows = []
    pm_rows = []

    machine_sheet = workbook["Machines List"]
    machine_iter = iter_non_empty_rows(machine_sheet)
    next(machine_iter, None)
    for row in machine_iter:
        equipment_rows.append(
            {
                "equipment_code": cstr(row[0]),
                "equipment_name": cstr(row[1]),
            }
        )

    spare_sheet = workbook["Spare Codes"]
    spare_iter = iter_non_empty_rows(spare_sheet)
    next(spare_iter, None)
    for row in spare_iter:
        spare_rows.append(
            {
                "spare_code": cstr(row[0]),
                "spare_name": cstr(row[1]) or cstr(row[0]),
                "machine_reference": cstr(row[2]),
            }
        )

    pm_sheet = workbook["Preventive Maintenance"]
    pm_iter = iter_non_empty_rows(pm_sheet)
    next(pm_iter, None)
    for row in pm_iter:
        pm_rows.append(
            {
                "frequency": cstr(row[0]),
                "equipment_type": cstr(row[1]),
                "equipment_code": cstr(row[2]),
                "check_point": cstr(row[3]),
                "method": cstr(row[4]),
                "criteria": cstr(row[5]),
            }
        )

    workbook.close()

    return {
        "equipment_rows": equipment_rows,
        "spare_rows": spare_rows,
        "pm_rows": pm_rows,
    }


def ensure_item_group(name):
    if frappe.db.exists("Item Group", name):
        return

    frappe.get_doc(
        {
            "doctype": "Item Group",
            "item_group_name": name,
            "parent_item_group": "All Item Groups",
            "is_group": 0,
        }
    ).insert(ignore_permissions=True)


def ensure_uom(uom_name):
    if frappe.db.exists("UOM", uom_name):
        return

    frappe.get_doc({"doctype": "UOM", "uom_name": uom_name}).insert(ignore_permissions=True)


def ensure_prerequisites():
    ensure_item_group(MAINTENANCE_SPARE_GROUP)
    ensure_uom(DEFAULT_SPARE_UOM)


def upsert_equipment(row):
    equipment_code = row["equipment_code"]
    equipment_name = row["equipment_name"] or equipment_code

    if frappe.db.exists("Maintenance Equipment", equipment_code):
        doc = frappe.get_doc("Maintenance Equipment", equipment_code)
        created = False
    else:
        doc = frappe.new_doc("Maintenance Equipment")
        doc.equipment_code = equipment_code
        created = True

    doc.machine_code = equipment_code
    doc.machine_name = equipment_name
    doc.active = 1
    doc.equipment_code = equipment_code
    doc.equipment_name = equipment_name
    doc.equipment_type = equipment_name
    doc.equipment_group = guess_equipment_group(equipment_code=equipment_code, equipment_name=equipment_name)
    doc.is_active = 1

    if created:
        doc.insert(ignore_permissions=True)
    else:
        doc.save(ignore_permissions=True)

    return created, doc.name


def upsert_spare_item(row):
    item_code = row["spare_code"]
    source_name = row["spare_name"] or item_code

    if frappe.db.exists("Item", item_code):
        item = frappe.get_doc("Item", item_code)
        created = False
    else:
        item = frappe.new_doc("Item")
        item.item_code = item_code
        item.item_group = MAINTENANCE_SPARE_GROUP
        item.stock_uom = DEFAULT_SPARE_UOM
        item.is_stock_item = 1
        item.disabled = 0
        created = True

    if created or not cstr(item.item_name) or cstr(item.item_name) == item_code:
        item.item_name = source_name

    if created:
        item.item_group = item.item_group or MAINTENANCE_SPARE_GROUP
        item.stock_uom = item.stock_uom or DEFAULT_SPARE_UOM
        item.is_stock_item = 1

    item.custom_is_maintenance_spare = 1

    if created:
        item.insert(ignore_permissions=True)
    else:
        item.save(ignore_permissions=True)

    return created, item.name


def resolve_equipment_link(machine_reference: str):
    machine_reference = cstr(machine_reference)
    if not machine_reference:
        return None

    if frappe.db.exists("Maintenance Equipment", machine_reference):
        return machine_reference

    exact_match = frappe.db.get_value(
        "Maintenance Equipment",
        {"machine_name": machine_reference},
        "name",
    )
    if exact_match:
        return exact_match

    exact_match = frappe.db.get_value(
        "Maintenance Equipment",
        {"equipment_name": machine_reference},
        "name",
    )
    if exact_match:
        return exact_match

    return None


def upsert_spare_mapping(item_code, row):
    machine_reference = cstr(row["machine_reference"])
    mapping_name = frappe.db.get_value(
        "Maintenance Spare Mapping",
        {"spare_item": item_code, "source_machine_name": machine_reference},
        "name",
    )

    if mapping_name:
        mapping = frappe.get_doc("Maintenance Spare Mapping", mapping_name)
        created = False
    else:
        mapping = frappe.new_doc("Maintenance Spare Mapping")
        created = True

    linked_equipment = resolve_equipment_link(machine_reference)
    mapping.machine = linked_equipment
    mapping.spare_item = item_code
    mapping.equipment = linked_equipment
    mapping.source_machine_name = machine_reference
    mapping.equipment_group = guess_equipment_group(source_reference=machine_reference)
    mapping.critical = int(frappe.db.get_value("Item", item_code, "custom_maintenance_critical") or 0)
    mapping.standard_qty = mapping.standard_qty or 1
    mapping.is_active = 1

    if created:
        mapping.insert(ignore_permissions=True)
    else:
        mapping.save(ignore_permissions=True)

    return created, bool(linked_equipment), bool(mapping.equipment_group)


def upsert_pm_plan(row):
    equipment = frappe.db.get_value("Maintenance Equipment", row["equipment_code"], "name")
    if not equipment:
        return False, False, row["equipment_code"]

    frequency_details = get_frequency_details(row["frequency"])
    frequency = frequency_details["label"]
    existing_name = frappe.db.get_value(
        "Preventive Maintenance Plan",
        {
            "equipment": equipment,
            "frequency": frequency,
            "activity_checklist": row["check_point"],
        },
        "name",
    )

    if existing_name:
        plan = frappe.get_doc("Preventive Maintenance Plan", existing_name)
        created = False
    else:
        plan = frappe.new_doc("Preventive Maintenance Plan")
        created = True

    plan.equipment = equipment
    plan.frequency = frequency
    plan.activity_checklist = row["check_point"]
    plan.method = row["method"]
    plan.criteria = row["criteria"]
    plan.responsible_role = plan.responsible_role or "Maintenance"
    plan.is_active = 1

    if created:
        plan.insert(ignore_permissions=True)
    else:
        plan.save(ignore_permissions=True)

    return created, True, None


def execute(workbook_path):
    ensure_prerequisites()
    data = read_workbook(workbook_path)

    result = {
        "equipment_created": 0,
        "equipment_updated": 0,
        "spare_items_created": 0,
        "spare_items_updated": 0,
        "spare_mappings_created": 0,
        "spare_mappings_updated": 0,
        "spare_mappings_linked_to_equipment": 0,
        "spare_mappings_group_tagged": 0,
        "pm_plans_created": 0,
        "pm_plans_updated": 0,
        "pm_rows_skipped_missing_equipment": [],
    }

    for row in data["equipment_rows"]:
        created, _ = upsert_equipment(row)
        result["equipment_created" if created else "equipment_updated"] += 1

    for row in data["spare_rows"]:
        item_created, item_code = upsert_spare_item(row)
        result["spare_items_created" if item_created else "spare_items_updated"] += 1

        mapping_created, linked_equipment, group_tagged = upsert_spare_mapping(item_code, row)
        result["spare_mappings_created" if mapping_created else "spare_mappings_updated"] += 1
        if linked_equipment:
            result["spare_mappings_linked_to_equipment"] += 1
        if group_tagged:
            result["spare_mappings_group_tagged"] += 1

    for row in data["pm_rows"]:
        created, imported, missing_equipment = upsert_pm_plan(row)
        if imported:
            result["pm_plans_created" if created else "pm_plans_updated"] += 1
        elif missing_equipment:
            result["pm_rows_skipped_missing_equipment"].append(missing_equipment)

    result["equipment_total"] = frappe.db.count("Maintenance Equipment")
    result["maintenance_spare_items_total"] = frappe.db.count("Item", {"custom_is_maintenance_spare": 1})
    result["spare_mappings_total"] = frappe.db.count("Maintenance Spare Mapping")
    result["pm_plans_total"] = frappe.db.count("Preventive Maintenance Plan")

    generated_dir().mkdir(parents=True, exist_ok=True)
    report_path = generated_dir() / "maintenance_inventory_import_result.json"
    report_path.write_text(json.dumps(result, indent=2), encoding="utf-8")

    frappe.db.commit()
    return result


def main():
    workbook_path = Path("/tmp/maintenance_inventory.xlsx")
    frappe.init(site="frontend", sites_path="/home/frappe/frappe-bench/sites")
    frappe.connect()
    try:
        print(json.dumps(execute(workbook_path), indent=2))
    finally:
        frappe.destroy()


if __name__ == "__main__":
    main()
