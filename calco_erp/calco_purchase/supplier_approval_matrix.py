from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import json

import frappe
from frappe import _
from frappe.custom.doctype.custom_field.custom_field import create_custom_fields
from frappe.desk.search import validate_and_sanitize_search_inputs
from frappe.model.document import Document
from frappe.utils import cint, flt, getdate, today

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


MATRIX_DOCTYPE = "Supplier Approval Matrix"
MATRIX_ITEM_FIELD = "item_code"
MATRIX_SUPPLIER_FIELD = "supplier"
MATRIX_TYPE_FIELD = "supplier_type"
MATRIX_STATUS_FIELD = "approval_status"
MATRIX_RATING_FIELD = "supplier_rating"
MATRIX_LEAD_TIME_FIELD = "lead_time"
MATRIX_PAYMENT_TERMS_FIELD = "payment_terms"
MATRIX_EFFECTIVE_DATE_FIELD = "effective_date"
MATRIX_EXPIRY_DATE_FIELD = "expiry_date"

RFQ_SUPPLIER_APPROVAL_STATUS_FIELD = "custom_approval_status"
RFQ_SUPPLIER_TYPE_FIELD = "custom_supplier_type"
RFQ_WARNING_FIELD = "custom_supplier_matrix_warning"
SQ_APPROVED_SUPPLIER_FIELD = "custom_approved_rfq_supplier"

STATUS_APPROVED = "Approved"
STATUS_CONDITIONAL = "Conditional Approval"
STATUS_BLOCKED = "Blocked"
STATUS_UNDER_EVALUATION = "Under Evaluation"
STATUS_EXPIRED = "Expired"

APPROVAL_PRIORITY = {
    STATUS_APPROVED: 0,
    STATUS_CONDITIONAL: 1,
    STATUS_EXPIRED: 2,
    STATUS_UNDER_EVALUATION: 3,
    STATUS_BLOCKED: 4,
}

WORKBOOK_HEADERS = {
    "item_code": "Product Code",
    "supplier_name": "Supplier Name",
    "basis_of_approval": "Basis of Approval",
    "payment_terms": "Payment Terms",
    "supplier_evaluation_form": "Supplier Evaluation Form",
    "supplier_rating": "Supplier Rating",
    "supplier_id": "Supplier ID",
    "supplier_agreement": "Supplier Agreement",
    "overseas": "Overseas",
    "expiry_date": "Expiration Date of ISO Certificate",
}

SUPPLIER_QUOTATION_CLIENT_SCRIPT_NAME = "Supplier Quotation Supplier Filter"


SUPPLIER_QUOTATION_CLIENT_SCRIPT = """
function supplierQuotationCollectValues(frm, fieldname) {
\treturn (frm.doc.items || [])
\t\t.map((row) => row[fieldname])
\t\t.filter(Boolean);
}

function setApprovedSupplierFieldProperties(frm, options) {
\tconst selectOptions = [""].concat(options || []).join("\\n");
\tfrm.set_df_property("supplier", "read_only", 1);
\tfrm.set_df_property("supplier", "hidden", 1);
\tfrm.set_df_property("supplier_name", "read_only", 1);
\tfrm.set_df_property("supplier_name", "hidden", 1);
\tfrm.set_df_property("custom_approved_rfq_supplier", "hidden", 0);
\tfrm.set_df_property("custom_approved_rfq_supplier", "reqd", options.length ? 1 : 0);
\tfrm.set_df_property("custom_approved_rfq_supplier", "options", selectOptions);
\tfrm.refresh_field("custom_approved_rfq_supplier");
}

function setSupplierQuotationSupplierQuery(frm) {
\tconst itemCodes = supplierQuotationCollectValues(frm, "item_code");
\tconst rfqNames = supplierQuotationCollectValues(frm, "request_for_quotation");

\tfrm.set_query("supplier", function() {
\t\treturn {
\t\t\tquery: "calco_erp.calco_purchase.supplier_approval_matrix.supplier_quotation_supplier_query",
\t\t\tfilters: {
\t\t\t\tdocname: frm.doc.name || "",
\t\t\t\titem_codes: JSON.stringify(itemCodes),
\t\t\t\trequest_for_quotations: JSON.stringify(rfqNames),
\t\t\t}
\t\t};
\t});
}

function applySupplierQuotationSupplierFilter(frm) {
\tsetSupplierQuotationSupplierQuery(frm);

\tfrappe.call({
\t\tmethod: "calco_erp.calco_purchase.supplier_approval_matrix.get_supplier_quotation_supplier_options",
\t\targs: {
\t\t\tdoc: frm.doc,
\t\t},
\t\tcallback: ({ message }) => {
\t\t\tconst info = message || {};
\t\t\tconst allowedSuppliers = info.allowed_suppliers || [];
\t\t\tfrm._supplierQuotationAllowedSuppliers = allowedSuppliers;
\t\t\tconsole.log("Supplier Query Called", allowedSuppliers);
\t\t\tsetApprovedSupplierFieldProperties(frm, allowedSuppliers);
\t\t\tif (allowedSuppliers.length === 1 && frm.doc.custom_approved_rfq_supplier !== allowedSuppliers[0]) {
\t\t\t\tfrm.set_value("custom_approved_rfq_supplier", allowedSuppliers[0]);
\t\t\t}
\t\t\tif (frm.doc.custom_approved_rfq_supplier && allowedSuppliers.length && !allowedSuppliers.includes(frm.doc.custom_approved_rfq_supplier)) {
\t\t\t\tfrm.set_value("custom_approved_rfq_supplier", "");
\t\t\t}
\t\t\tif (frm.doc.custom_approved_rfq_supplier) {
\t\t\t\tfrm.set_value("supplier", frm.doc.custom_approved_rfq_supplier);
\t\t\t}
\t\t\tif (frm.doc.supplier && allowedSuppliers.length && !allowedSuppliers.includes(frm.doc.supplier)) {
\t\t\t\tfrm.set_value("supplier", "");
\t\t\t}
\t\t\tfrm.refresh_field("supplier");
\t\t},
\t});
}

frappe.ui.form.on("Supplier Quotation", {
\tsetup(frm) {
\t\tif (!frm.__supplierQueryRealtimeBound) {
\t\t\tfrm.__supplierQueryRealtimeBound = true;
\t\t\tfrappe.realtime.on("supplier_filter_debug", (payload) => {
\t\t\t\tconsole.log("Custom Supplier Query Executed", payload);
\t\t\t\tfrappe.show_alert({
\t\t\t\t\tmessage: __("Custom Supplier Query Executed"),
\t\t\t\t\tindicator: "green",
\t\t\t\t});
\t\t\t});
\t\t}
\t\tsetSupplierQuotationSupplierQuery(frm);
\t\tfrappe.show_alert({
\t\t\tmessage: __("Supplier Quotation Client Script Loaded"),
\t\t\tindicator: "green",
\t\t});
\t},
\tonload(frm) {
\t\tapplySupplierQuotationSupplierFilter(frm);
\t},
\trefresh(frm) {
\t\tapplySupplierQuotationSupplierFilter(frm);
\t},
\tcustom_approved_rfq_supplier(frm) {
\t\tconst allowed = frm._supplierQuotationAllowedSuppliers || [];
\t\tif (frm.doc.custom_approved_rfq_supplier && allowed.length && !allowed.includes(frm.doc.custom_approved_rfq_supplier)) {
\t\t\tfrappe.msgprint(__("Selected approved RFQ supplier is not allowed for this Supplier Quotation."));
\t\t\tfrm.set_value("custom_approved_rfq_supplier", "");
\t\t\tfrm.set_value("supplier", "");
\t\t\treturn;
\t\t}
\t\tfrm.set_value("supplier", frm.doc.custom_approved_rfq_supplier || "");
\t},
\tsupplier(frm) {
\t\tconst allowed = frm._supplierQuotationAllowedSuppliers || [];
\t\tif (frm.doc.supplier && allowed.length && !allowed.includes(frm.doc.supplier)) {
\t\t\tfrappe.msgprint(__("Selected supplier is not allowed for this Supplier Quotation."));
\t\t\tfrm.set_value("supplier", "");
\t\t\tfrm.set_value("custom_approved_rfq_supplier", "");
\t\t}
\t},
});

frappe.ui.form.on("Supplier Quotation Item", {
\titems_add(frm) {
\t\tapplySupplierQuotationSupplierFilter(frm);
\t},
\titems_remove(frm) {
\t\tapplySupplierQuotationSupplierFilter(frm);
\t},
\titem_code(frm) {
\t\tapplySupplierQuotationSupplierFilter(frm);
\t},
\trequest_for_quotation(frm) {
\t\tapplySupplierQuotationSupplierFilter(frm);
\t},
});
""".strip()


def ensure_supplier_approval_setup():
    create_custom_fields(
        {
            "Request for Quotation": [
                {
                    "fieldname": RFQ_WARNING_FIELD,
                    "fieldtype": "Small Text",
                    "label": "Supplier Matrix Warning",
                    "insert_after": "suppliers",
                    "read_only": 1,
                }
            ],
            "Request for Quotation Supplier": [
                {
                    "fieldname": RFQ_SUPPLIER_APPROVAL_STATUS_FIELD,
                    "fieldtype": "Data",
                    "label": "Approval Status",
                    "insert_after": "supplier_name",
                    "read_only": 1,
                    "in_list_view": 1,
                },
                {
                    "fieldname": RFQ_SUPPLIER_TYPE_FIELD,
                    "fieldtype": "Data",
                    "label": "Supplier Type",
                    "insert_after": RFQ_SUPPLIER_APPROVAL_STATUS_FIELD,
                    "read_only": 1,
                },
            ],
            "Supplier Quotation": [
                {
                    "fieldname": SQ_APPROVED_SUPPLIER_FIELD,
                    "fieldtype": "Select",
                    "label": "Approved RFQ Supplier",
                    "insert_after": "supplier",
                    "description": "Select only from approved suppliers derived from the linked RFQ or supplier approval matrix.",
                }
            ],
        },
        update=True,
        ignore_validate=True,
    )
    ensure_supplier_quotation_client_script()
    frappe.clear_cache()


def ensure_supplier_quotation_client_script():
    if not frappe.db.exists("DocType", "Client Script"):
        return

    existing_name = frappe.db.get_value(
        "Client Script",
        {"dt": "Supplier Quotation", "name": SUPPLIER_QUOTATION_CLIENT_SCRIPT_NAME},
        "name",
    )
    if not existing_name:
        existing_name = frappe.db.get_value(
            "Client Script",
            {"dt": "Supplier Quotation", "enabled": 1},
            "name",
        )

    values = {
        "dt": "Supplier Quotation",
        "enabled": 1,
        "view": "Form",
        "script": SUPPLIER_QUOTATION_CLIENT_SCRIPT,
    }

    if existing_name:
        doc = frappe.get_doc("Client Script", existing_name)
        for key, value in values.items():
            doc.set(key, value)
        if doc.name != SUPPLIER_QUOTATION_CLIENT_SCRIPT_NAME:
            doc.name = SUPPLIER_QUOTATION_CLIENT_SCRIPT_NAME
        doc.save(ignore_permissions=True)
    else:
        doc = frappe.get_doc(
            {
                "doctype": "Client Script",
                "name": SUPPLIER_QUOTATION_CLIENT_SCRIPT_NAME,
                **values,
            }
        )
        doc.insert(ignore_permissions=True)


def normalize_item_code(value: str | None) -> str:
    return (value or "").strip().upper()


def normalize_supplier_name(value: str | None) -> str:
    supplier_name = (value or "").strip()
    if supplier_name.lower().startswith("blocked "):
        supplier_name = supplier_name[8:].strip()
    return supplier_name


def condensed_token(value: str | None) -> str:
    return "".join(ch for ch in (value or "").upper() if ch.isalnum())


def derive_supplier_type(row: dict[str, object]) -> str:
    overseas = (row.get("overseas") or "").strip().lower()
    return "Overseas" if overseas in {"yes", "y", "1", "true"} else "Local"


def parse_workbook_date(value) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    try:
        return getdate(value).isoformat()
    except Exception:
        return None


def derive_approval_status(row: dict[str, object], expiry_date: str | None) -> str:
    supplier_name = (row.get("supplier_name") or "").strip()
    basis = (row.get("basis_of_approval") or "").strip()
    agreement = (row.get("supplier_agreement") or "").strip()
    evaluation = (row.get("supplier_evaluation_form") or "").strip()
    status_text = " ".join(filter(None, [supplier_name, basis, agreement, evaluation])).lower()

    if supplier_name.lower().startswith("blocked ") or "blocked" in agreement.lower():
        return STATUS_BLOCKED
    # The workbook's expiry column is the supplier ISO certificate expiry, not the
    # approval matrix lifecycle state. Keep it stored on the matrix row, but do not
    # downgrade otherwise-approved suppliers solely because the certificate date is old.
    if "expired" in status_text:
        return STATUS_EXPIRED
    if "conditional" in status_text:
        return STATUS_CONDITIONAL
    if basis or agreement or evaluation:
        return STATUS_APPROVED
    return STATUS_UNDER_EVALUATION


def get_or_create_supplier(supplier_name: str, supplier_type: str = "Company") -> str:
    existing_name = frappe.db.get_value("Supplier", {"supplier_name": supplier_name}, "name")
    if existing_name:
        return existing_name
    if frappe.db.exists("Supplier", supplier_name):
        return supplier_name

    first_token = (supplier_name.split() or [""])[0]
    token_key = condensed_token(first_token)
    if token_key and frappe.db.exists("Supplier", token_key):
        return token_key

    for row in frappe.get_all("Supplier", fields=["name", "supplier_name"], limit_page_length=0):
        if condensed_token(row.get("supplier_name")) == condensed_token(supplier_name):
            return row["name"]
        if token_key and condensed_token(row.get("name")) == token_key:
            return row["name"]

    supplier = frappe.get_doc(
        {
            "doctype": "Supplier",
            "supplier_name": supplier_name,
            "supplier_type": supplier_type or "Company",
            "supplier_group": "All Supplier Groups" if frappe.db.exists("Supplier Group", "All Supplier Groups") else None,
        }
    )
    supplier.insert(ignore_permissions=True)
    return supplier.name


def get_supplier_display_name(supplier: str | None) -> str:
    if not supplier:
        return ""
    return (
        frappe.db.get_value("Supplier", supplier, "supplier_name")
        or frappe.db.get_value("Supplier", {"supplier_name": supplier}, "supplier_name")
        or supplier
    )


def find_existing_matrix_name(item_code: str, supplier: str, source_supplier_name: str) -> str | None:
    existing_name = frappe.db.get_value(
        MATRIX_DOCTYPE,
        {
            MATRIX_ITEM_FIELD: item_code,
            MATRIX_SUPPLIER_FIELD: supplier,
        },
        "name",
    )
    if existing_name:
        return existing_name

    target_keys = {
        condensed_token(source_supplier_name),
        condensed_token(supplier),
        condensed_token(get_supplier_display_name(supplier)),
    }
    rows = frappe.get_all(
        MATRIX_DOCTYPE,
        filters={MATRIX_ITEM_FIELD: item_code},
        fields=["name", MATRIX_SUPPLIER_FIELD],
        limit_page_length=0,
    )
    for row in rows:
        supplier_name = row.get(MATRIX_SUPPLIER_FIELD)
        candidate_keys = {
            condensed_token(supplier_name),
            condensed_token(get_supplier_display_name(supplier_name)),
        }
        if target_keys.intersection(candidate_keys):
            return row["name"]
    return None


def load_supplier_master_rows(file_path: str) -> list[dict[str, object]]:
    if not load_workbook:
        frappe.throw(_("openpyxl is required to import the Supplier Master List."))

    workbook_path = Path(file_path)
    if not workbook_path.exists():
        frappe.throw(_("Supplier Master List file not found at {0}.").format(file_path))

    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook.worksheets[0]
    headers = [str(cell).strip() if cell is not None else "" for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {header: idx for idx, header in enumerate(headers)}

    rows = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        row = {}
        for key, header in WORKBOOK_HEADERS.items():
            idx = header_index.get(header)
            row[key] = values[idx] if idx is not None and idx < len(values) else None
        rows.append(row)
    return rows


@frappe.whitelist()
def import_supplier_approval_matrix(file_path: str, create_missing_suppliers: int = 1) -> dict[str, object]:
    rows = load_supplier_master_rows(file_path)
    create_missing = cint(create_missing_suppliers)
    summary = Counter()
    by_item_supplier: dict[tuple[str, str], dict[str, object]] = {}

    for source_row in rows:
        item_code = normalize_item_code(source_row.get("item_code"))
        supplier_name = normalize_supplier_name(source_row.get("supplier_name"))
        if not item_code or not supplier_name:
            continue
        if not frappe.db.exists("Item", item_code):
            summary["skipped_missing_item"] += 1
            continue

        expiry_date = parse_workbook_date(source_row.get("expiry_date"))
        status = derive_approval_status(source_row, expiry_date)
        supplier_type = derive_supplier_type(source_row)
        payment_terms = str(source_row.get("payment_terms") or "").strip()
        supplier_rating = str(source_row.get("supplier_rating") or "").strip()
        lead_time = ""
        effective_date = today()

        if create_missing:
            supplier = get_or_create_supplier(supplier_name)
        else:
            supplier = frappe.db.get_value("Supplier", {"supplier_name": supplier_name}, "name")
            if not supplier:
                summary["skipped_missing_supplier"] += 1
                continue

        key = (item_code, supplier)
        existing = by_item_supplier.get(key)
        candidate = {
            MATRIX_ITEM_FIELD: item_code,
            MATRIX_SUPPLIER_FIELD: supplier,
            MATRIX_TYPE_FIELD: supplier_type,
            MATRIX_STATUS_FIELD: status,
            MATRIX_RATING_FIELD: supplier_rating,
            MATRIX_LEAD_TIME_FIELD: lead_time,
            MATRIX_PAYMENT_TERMS_FIELD: payment_terms,
            MATRIX_EFFECTIVE_DATE_FIELD: effective_date,
            MATRIX_EXPIRY_DATE_FIELD: expiry_date,
            "_source_supplier_name": supplier_name,
        }
        if not existing or APPROVAL_PRIORITY[status] < APPROVAL_PRIORITY[existing[MATRIX_STATUS_FIELD]]:
            by_item_supplier[key] = candidate

    for values in by_item_supplier.values():
        existing_name = find_existing_matrix_name(
            values[MATRIX_ITEM_FIELD],
            values[MATRIX_SUPPLIER_FIELD],
            values.get("_source_supplier_name") or get_supplier_display_name(values[MATRIX_SUPPLIER_FIELD]),
        )
        values.pop("_source_supplier_name", None)
        if existing_name:
            doc = frappe.get_doc(MATRIX_DOCTYPE, existing_name)
            doc.update(values)
            doc.save(ignore_permissions=True)
            summary["updated"] += 1
        else:
            doc = frappe.get_doc({"doctype": MATRIX_DOCTYPE, **values})
            doc.insert(ignore_permissions=True)
            summary["created"] += 1
        summary[values[MATRIX_STATUS_FIELD]] += 1

    return {
        "file_path": file_path,
        "imported_rows": len(by_item_supplier),
        "summary": dict(summary),
    }


def get_matrix_rows_for_item(item_code: str) -> list[dict]:
    normalized = normalize_item_code(item_code)
    if not normalized:
        return []
    return frappe.get_all(
        MATRIX_DOCTYPE,
        filters={MATRIX_ITEM_FIELD: normalized},
        fields=[
            "name",
            MATRIX_ITEM_FIELD,
            MATRIX_SUPPLIER_FIELD,
            MATRIX_TYPE_FIELD,
            MATRIX_STATUS_FIELD,
            MATRIX_RATING_FIELD,
            MATRIX_LEAD_TIME_FIELD,
            MATRIX_PAYMENT_TERMS_FIELD,
            MATRIX_EFFECTIVE_DATE_FIELD,
            MATRIX_EXPIRY_DATE_FIELD,
        ],
        limit_page_length=0,
    )


def get_rfq_supplier_candidates(item_codes: list[str]) -> dict[str, object]:
    normalized_items = unique_non_empty(normalize_item_code(code) for code in item_codes)
    item_rows = {item_code: get_matrix_rows_for_item(item_code) for item_code in normalized_items}
    missing_matrix_items = [item_code for item_code, rows in item_rows.items() if not rows]
    if missing_matrix_items:
        frappe.throw(
            _("No supplier approval matrix exists for item(s): {0}. RFQ creation is blocked.").format(
                ", ".join(missing_matrix_items)
            )
        )

    missing_approved_items = []
    eligible_by_item: dict[str, dict[str, dict]] = {}
    for item_code, rows in item_rows.items():
        approved_rows = [row for row in rows if row.get(MATRIX_STATUS_FIELD) == STATUS_APPROVED]
        if not approved_rows:
            missing_approved_items.append(item_code)

        eligible = {}
        for row in rows:
            status = row.get(MATRIX_STATUS_FIELD)
            if status not in {STATUS_APPROVED, STATUS_CONDITIONAL}:
                continue
            eligible[row[MATRIX_SUPPLIER_FIELD]] = row
        eligible_by_item[item_code] = eligible

    if missing_approved_items:
        frappe.throw(
            _("No approved supplier exists for item(s): {0}. RFQ creation is blocked.").format(
                ", ".join(missing_approved_items)
            )
        )

    common_supplier_names = None
    for eligible in eligible_by_item.values():
        names = set(eligible.keys())
        common_supplier_names = names if common_supplier_names is None else common_supplier_names.intersection(names)

    common_supplier_names = common_supplier_names or set()
    if not common_supplier_names:
        frappe.throw(
            _("No common approved or conditional supplier exists across all selected items. Split the RFQ by item or supplier.")
        )

    candidates = []
    warnings = []
    for supplier in sorted(common_supplier_names):
        per_item_rows = [eligible_by_item[item_code][supplier] for item_code in normalized_items]
        statuses = {row.get(MATRIX_STATUS_FIELD) for row in per_item_rows}
        aggregate_status = STATUS_CONDITIONAL if STATUS_CONDITIONAL in statuses else STATUS_APPROVED
        supplier_type = next((row.get(MATRIX_TYPE_FIELD) for row in per_item_rows if row.get(MATRIX_TYPE_FIELD)), "")
        payment_terms = next((row.get(MATRIX_PAYMENT_TERMS_FIELD) for row in per_item_rows if row.get(MATRIX_PAYMENT_TERMS_FIELD)), "")
        rating = next((row.get(MATRIX_RATING_FIELD) for row in per_item_rows if row.get(MATRIX_RATING_FIELD)), "")
        if aggregate_status == STATUS_CONDITIONAL:
            warnings.append(supplier)
        candidates.append(
            {
                "supplier": supplier,
                "approval_status": aggregate_status,
                "supplier_type": supplier_type,
                "payment_terms": payment_terms,
                "supplier_rating": rating,
            }
        )

    candidates.sort(key=lambda row: (APPROVAL_PRIORITY[row["approval_status"]], row["supplier"]))
    return {
        "item_codes": normalized_items,
        "candidates": candidates,
        "conditional_suppliers": warnings,
    }


def set_rfq_supplier_rows(doc: Document, candidates: list[dict], warning_message: str = ""):
    doc.set("suppliers", [])
    for row in candidates:
        child = doc.append(
            "suppliers",
            {
                "supplier": row["supplier"],
                RFQ_SUPPLIER_APPROVAL_STATUS_FIELD: row["approval_status"],
                RFQ_SUPPLIER_TYPE_FIELD: row["supplier_type"],
            },
        )
        if hasattr(child, "supplier_name") and not child.get("supplier_name"):
            child.supplier_name = frappe.db.get_value("Supplier", row["supplier"], "supplier_name")

    if doc.meta.has_field(RFQ_WARNING_FIELD):
        doc.set(RFQ_WARNING_FIELD, warning_message)


def build_rfq_warning_message(candidates: list[dict], conditional_suppliers: list[str]) -> str:
    if conditional_suppliers:
        return _("Conditional Approval supplier(s) included: {0}. Proceed with caution.").format(
            ", ".join(conditional_suppliers)
        )
    return ""


def apply_supplier_matrix_to_rfq(doc: Document):
    if cint(getattr(doc, "docstatus", 0)) == 1:
        return
    item_codes = [row.get("item_code") for row in doc.get("items", []) if row.get("item_code")]
    if not item_codes:
        return
    result = get_rfq_supplier_candidates(item_codes)
    warning_message = build_rfq_warning_message(result["candidates"], result["conditional_suppliers"])
    set_rfq_supplier_rows(doc, result["candidates"], warning_message=warning_message)


def validate_request_for_quotation_supplier_matrix(doc, method=None):
    if doc.doctype != "Request for Quotation":
        return
    apply_supplier_matrix_to_rfq(doc)


@frappe.whitelist()
def make_request_for_quotation_with_supplier_matrix(source_name, target_doc=None):
    from erpnext.stock.doctype.material_request.material_request import make_request_for_quotation

    doc = make_request_for_quotation(source_name, target_doc=target_doc)
    apply_supplier_matrix_to_rfq(doc)
    return doc


def unique_non_empty(values) -> list[str]:
    seen = []
    for value in values:
        if not value or value in seen:
            continue
        seen.append(value)
    return seen


def _coerce_doc(doc) -> frappe._dict:
    if isinstance(doc, str):
        doc = frappe.parse_json(doc)
    return frappe._dict(doc or {})


def _coerce_rows(rows) -> list[frappe._dict]:
    return [frappe._dict(row or {}) for row in (rows or [])]


def _parse_filter_list(value) -> list[str]:
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        try:
            parsed = frappe.parse_json(stripped)
            if isinstance(parsed, list):
                return unique_non_empty(str(v).strip() for v in parsed if v)
        except Exception:
            pass
        return unique_non_empty(part.strip() for part in stripped.split(",") if part.strip())
    if isinstance(value, (list, tuple, set)):
        return unique_non_empty(str(v).strip() for v in value if v)
    return []


def get_allowed_supplier_rows_for_supplier_quotation_filters(filters) -> dict[str, object]:
    filters = frappe._dict(filters or {})
    rfq_names = _parse_filter_list(filters.get("request_for_quotations"))
    item_codes = unique_non_empty(normalize_item_code(code) for code in _parse_filter_list(filters.get("item_codes")))

    if rfq_names:
        rfq_supplier_sets: list[set[str]] = []
        allowed_rows: dict[str, dict] = {}
        for rfq_name in rfq_names:
            rfq_suppliers = frappe.get_all(
                "Request for Quotation Supplier",
                filters={"parent": rfq_name},
                fields=[
                    "supplier",
                    "supplier_name",
                    RFQ_SUPPLIER_APPROVAL_STATUS_FIELD,
                    RFQ_SUPPLIER_TYPE_FIELD,
                ],
                limit_page_length=0,
            )
            supplier_set = {row.get("supplier") for row in rfq_suppliers if row.get("supplier")}
            rfq_supplier_sets.append(supplier_set)
            for row in rfq_suppliers:
                supplier = row.get("supplier")
                if not supplier:
                    continue
                allowed_rows[supplier] = {
                    "supplier": supplier,
                    "supplier_name": row.get("supplier_name") or get_supplier_display_name(supplier),
                    "approval_status": row.get(RFQ_SUPPLIER_APPROVAL_STATUS_FIELD) or "",
                    "supplier_type": row.get(RFQ_SUPPLIER_TYPE_FIELD) or "",
                }

        allowed_suppliers = set.intersection(*rfq_supplier_sets) if rfq_supplier_sets else set()
        return {
            "source": "request_for_quotation",
            "request_for_quotations": rfq_names,
            "item_codes": item_codes,
            "allowed_suppliers": sorted(allowed_suppliers),
            "rows": [allowed_rows[name] for name in sorted(allowed_suppliers) if name in allowed_rows],
        }

    if not item_codes:
        return {
            "source": "none",
            "request_for_quotations": [],
            "item_codes": [],
            "allowed_suppliers": [],
            "rows": [],
        }

    result = get_rfq_supplier_candidates(item_codes)
    return {
        "source": "supplier_approval_matrix",
        "request_for_quotations": [],
        "item_codes": result["item_codes"],
        "allowed_suppliers": [row["supplier"] for row in result["candidates"]],
        "rows": result["candidates"],
    }


def get_allowed_supplier_rows_for_supplier_quotation(doc) -> dict[str, object]:
    doc = _coerce_doc(doc)
    item_rows = _coerce_rows(doc.get("items"))
    return get_allowed_supplier_rows_for_supplier_quotation_filters(
        {
            "request_for_quotations": unique_non_empty(row.get("request_for_quotation") for row in item_rows),
            "item_codes": unique_non_empty(row.get("item_code") for row in item_rows),
        }
    )


@frappe.whitelist()
def get_supplier_quotation_supplier_options(doc=None):
    return get_allowed_supplier_rows_for_supplier_quotation(doc)


@frappe.whitelist()
@validate_and_sanitize_search_inputs
def supplier_quotation_supplier_query(doctype, txt, searchfield, start, page_len, filters):
    filters = frappe._dict(filters or {})
    info = get_allowed_supplier_rows_for_supplier_quotation_filters(filters)
    allowed_suppliers = info.get("allowed_suppliers") or []
    log_payload = {
        "timestamp": frappe.utils.now(),
        "user": frappe.session.user,
        "rfq": info.get("request_for_quotations") or _parse_filter_list(filters.get("request_for_quotations")),
        "item_codes": info.get("item_codes") or _parse_filter_list(filters.get("item_codes")),
        "allowed_suppliers": allowed_suppliers,
        "docname": filters.get("docname"),
        "txt": txt,
    }
    frappe.logger("supplier_filter").info(json.dumps(log_payload, default=str))
    try:
        frappe.publish_realtime(
            "supplier_filter_debug",
            {"message": "Custom Supplier Query Executed", **log_payload},
            user=frappe.session.user,
            after_commit=False,
        )
    except Exception:
        pass
    if not allowed_suppliers:
        return []

    txt = (txt or "").strip().lower()
    rows = []
    for supplier in allowed_suppliers:
        supplier_name = get_supplier_display_name(supplier)
        haystacks = {(supplier or "").lower(), (supplier_name or "").lower()}
        if txt and not any(txt in haystack for haystack in haystacks):
            continue
        rows.append([supplier, supplier_name])

    return rows[start : start + page_len]


def validate_supplier_quotation_supplier_matrix(doc, method=None):
    if doc.doctype != "Supplier Quotation":
        return

    info = get_allowed_supplier_rows_for_supplier_quotation(doc.as_dict())
    allowed_suppliers = set(info.get("allowed_suppliers") or [])
    approved_supplier = doc.get(SQ_APPROVED_SUPPLIER_FIELD)
    original_supplier = doc.get("supplier")
    if approved_supplier and original_supplier and original_supplier != approved_supplier:
        frappe.throw(
            _("Supplier must match Approved RFQ Supplier: {0}.").format(approved_supplier)
        )
    if approved_supplier:
        doc.supplier = approved_supplier
    supplier = doc.get("supplier")
    if not supplier or not allowed_suppliers:
        return

    if supplier not in allowed_suppliers:
        if info.get("source") == "request_for_quotation":
            frappe.throw(
                _(
                    "Supplier {0} is not one of the supplier(s) selected in linked Request for Quotation: {1}."
                ).format(
                    supplier,
                    ", ".join(info.get("request_for_quotations") or []),
                )
            )

        frappe.throw(
            _(
                "Supplier {0} is not approved in the Supplier Approval Matrix for item(s): {1}."
            ).format(supplier, ", ".join(info.get("item_codes") or []))
        )
