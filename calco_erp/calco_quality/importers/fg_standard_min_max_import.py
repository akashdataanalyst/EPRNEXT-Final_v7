from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import frappe
from openpyxl import load_workbook

from calco_erp.calco_quality.fg_quality_setup import (
    FG_MANUAL_REVIEW_PARAMETERS,
    canonicalize_fg_testing_method_name,
    ensure_fg_quality_setup,
    get_true_fg_testing_method_name_set,
)
from calco_erp.calco_quality.rm_testing_utils import normalize_text, parse_float


ITEM_CODE_INDEX = 1
FG_STANDARD_COMMIT_BATCH_SIZE = 500
MAX_STORABLE_FG_LIMIT = 999999999999.0
HEADER_MIN_PATTERN = re.compile(r"\bmin(?:imum)?\b", flags=re.IGNORECASE)
HEADER_MAX_PATTERN = re.compile(r"\bmax(?:imum)?\b", flags=re.IGNORECASE)
IGNORED_HEADER_PATTERNS = (
    re.compile(r"\btest condition\b", flags=re.IGNORECASE),
)

DEFAULT_PARAMETER_UNITS = {
    "MFI": "gm/10min",
    "Density": "gm/cc",
    "Ash Content": "%",
    "Moisture Content": "%",
    "Tensile Strength at Yield": "MPa",
    "Young Modulus": "MPa",
    "Elongation at Yield": "%",
    "Flexural Strength": "MPa",
    "Flexural Modulus": "MPa",
    "Notch Izod Impact": "J/m",
    "Shrinkage - Linear": "%",
    "Shrinkage - Transverse": "%",
    "Glow Wire Test": "Deg C",
    "GWT @0.8": "Deg C",
    "GWT @1.6": "Deg C",
    "HDT (Deg C)": "Deg C",
    "Water Absorption": "%",
    "Melting Point": "Deg C",
    "Vicat Softening Point": "Deg C",
    "Coefficient of Linear Thermal Expansion": "um/m Deg C",
    "Dielectric Strength": "KV/mm",
    "Volume Resistivity": "Ohm.cm",
    "Surface Resistivity": "Ohm",
    "Delta E": "Delta E",
    "Transmittance": "%",
    "GWT or UL Seconds": "sec",
}

SINGLE_VALUE_FIELD_BY_PARAMETER = {
    "Color Variation (Black Spot Free)": "target_value",
    "CTI (V)": "minimum_value",
    "Delta E": "maximum_value",
    "Dielectric Constant": "target_value",
    "Dielectric Strength": "minimum_value",
    "Finish (Shine/Dull)": "target_value",
    "Glow Wire Test": "minimum_value",
    "GWT @0.8": "minimum_value",
    "GWT @1.6": "minimum_value",
    "GWT or UL Seconds": "minimum_value",
    "Long Cut": "target_value",
    "Metal Contamination": "maximum_value",
    "Moisture Content": "maximum_value",
    "Surface Resistivity": "minimum_value",
    "Transmittance": "minimum_value",
    "UL 94 Burning Test @0.8": "target_value",
    "UL 94 Burning Test @1.6": "target_value",
    "UL94 Burning Testing": "target_value",
    "Un-Notched Izod Impact Strength": "minimum_value",
    "Un-annealed, HDT": "minimum_value",
    "Vicat Softening Point": "minimum_value",
    "Volume Resistivity": "minimum_value",
    "Water Absorption": "maximum_value",
}

HEADER_PARAMETER_PATTERNS = [
    (re.compile(r"\bmfi\b", flags=re.IGNORECASE), "MFI"),
    (re.compile(r"\bdensity\b", flags=re.IGNORECASE), "Density"),
    (re.compile(r"\bash(?: content)?\b", flags=re.IGNORECASE), "Ash Content"),
    (re.compile(r"\bmoisture(?: content)?\b", flags=re.IGNORECASE), "Moisture Content"),
    (
        re.compile(r"(\bts\b.*\byield\b)|(\btensile strength at yield\b)", flags=re.IGNORECASE),
        "Tensile Strength at Yield",
    ),
    (
        re.compile(r"(\btensile modulus\b)|(\byoung modulus\b)", flags=re.IGNORECASE),
        "Young Modulus",
    ),
    (
        re.compile(r"(\bel\s*@?\s*yl\b)|(\belongation at yield\b)", flags=re.IGNORECASE),
        "Elongation at Yield",
    ),
    (
        re.compile(r"(\bflexural modulus\b)|(\bflexural ms\b)", flags=re.IGNORECASE),
        "Flexural Modulus",
    ),
    (
        re.compile(r"(\bflexural strength\b)|(\bfs @ pl\b)|(\bstrength @ pl\b)", flags=re.IGNORECASE),
        "Flexural Strength",
    ),
    (
        re.compile(r"\bun[- ]?notched izod\b", flags=re.IGNORECASE),
        "Un-Notched Izod Impact Strength",
    ),
    (re.compile(r"\bizod imp\b", flags=re.IGNORECASE), "Notch Izod Impact"),
    (
        re.compile(r"\bflammability\b.*\b0\.8\b", flags=re.IGNORECASE),
        "UL 94 Burning Test @0.8",
    ),
    (
        re.compile(r"\bflammability\b.*\b1\.6\b", flags=re.IGNORECASE),
        "UL 94 Burning Test @1.6",
    ),
    (
        re.compile(r"\bflammability\b.*\b3\.2\b", flags=re.IGNORECASE),
        "UL94 Burning Testing",
    ),
    (
        re.compile(r"\bglow wire test\b.*\b3\.2\b", flags=re.IGNORECASE),
        "Glow Wire Test",
    ),
    (re.compile(r"\bgwt\b.*\b0\.8\b", flags=re.IGNORECASE), "GWT @0.8"),
    (re.compile(r"\bgwt\b.*\b1\.6\b", flags=re.IGNORECASE), "GWT @1.6"),
    (re.compile(r"\bgwt or ul seconds\b", flags=re.IGNORECASE), "GWT or UL Seconds"),
    (re.compile(r"\bun[- ]?annealed\b.*\bhdt\b", flags=re.IGNORECASE), "Un-annealed, HDT"),
    (re.compile(r"\bhdt\b", flags=re.IGNORECASE), "HDT (Deg C)"),
    (
        re.compile(r"\bmould shrinkage\b.*\blinear\b", flags=re.IGNORECASE),
        "Shrinkage - Linear",
    ),
    (
        re.compile(r"\bmould shrinkage\b.*\btransverse\b", flags=re.IGNORECASE),
        "Shrinkage - Transverse",
    ),
    (re.compile(r"\bwater absorption\b", flags=re.IGNORECASE), "Water Absorption"),
    (re.compile(r"\bmelting(?: point)?\b", flags=re.IGNORECASE), "Melting Point"),
    (
        re.compile(r"\bvicat softening point\b", flags=re.IGNORECASE),
        "Vicat Softening Point",
    ),
    (
        re.compile(r"\bcoefficient of linear thermal expansion\b", flags=re.IGNORECASE),
        "Coefficient of Linear Thermal Expansion",
    ),
    (
        re.compile(r"(\bcomparative tracking index\b)|(\bcti\b)", flags=re.IGNORECASE),
        "CTI (V)",
    ),
    (re.compile(r"\bdielectric strength\b", flags=re.IGNORECASE), "Dielectric Strength"),
    (re.compile(r"\bdielectric constant\b", flags=re.IGNORECASE), "Dielectric Constant"),
    (re.compile(r"\bvolume resistivity\b", flags=re.IGNORECASE), "Volume Resistivity"),
    (re.compile(r"\bsurface resistivity\b", flags=re.IGNORECASE), "Surface Resistivity"),
    (re.compile(r"\blong cut\b", flags=re.IGNORECASE), "Long Cut"),
    (
        re.compile(r"(\bcolor variation\b)|(\bblack spot free\b)", flags=re.IGNORECASE),
        "Color Variation (Black Spot Free)",
    ),
    (re.compile(r"\bfinish\b", flags=re.IGNORECASE), "Finish (Shine/Dull)"),
    (re.compile(r"\bdelta e\b", flags=re.IGNORECASE), "Delta E"),
    (re.compile(r"\btransmittance\b", flags=re.IGNORECASE), "Transmittance"),
    (re.compile(r"\bmetal contamination\b", flags=re.IGNORECASE), "Metal Contamination"),
]


def import_workbook(path: str) -> dict[str, object]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        frappe.throw(f"Workbook not found at {workbook_path}")

    ensure_fg_quality_setup()
    allowed_parameters = get_true_fg_testing_method_name_set()
    rows = load_rows(workbook_path)
    if not rows:
        return {
            "grades_updated": 0,
            "parameter_columns": 0,
            "rows_updated": 0,
        }

    parameter_columns, ignored_headers = derive_parameter_columns(rows[0], allowed_parameters)
    item_standards, unmatched_items, repeated_items = build_item_standard_map(rows[1:], parameter_columns)
    sync_summary = sync_fg_parameter_standards(item_standards)

    frappe.clear_cache()
    frappe.db.commit()
    return {
        "parameter_columns": len(parameter_columns),
        "ignored_headers": ignored_headers,
        "unmatched_items": unmatched_items,
        "repeated_source_items": repeated_items,
        **sync_summary,
    }


def load_rows(workbook_path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def derive_parameter_columns(header_row, allowed_parameters: set[str]) -> tuple[list[dict[str, object]], list[str]]:
    columns: list[dict[str, object]] = []
    ignored_headers: list[str] = []
    for index, value in enumerate(header_row):
        if index <= ITEM_CODE_INDEX:
            continue

        raw_header = normalize_text(value)
        if not raw_header:
            continue

        parsed = parse_standard_header(raw_header, allowed_parameters)
        if not parsed:
            ignored_headers.append(raw_header)
            continue

        columns.append(
            {
                "index": index,
                "header": raw_header,
                **parsed,
            }
        )
    return columns, ignored_headers


def parse_standard_header(raw_header: str, allowed_parameters: set[str]) -> dict[str, object] | None:
    if any(pattern.search(raw_header) for pattern in IGNORED_HEADER_PATTERNS):
        return None

    fieldname = get_limit_fieldname(raw_header)
    parameter_label = strip_limit_tokens(raw_header)
    parameter_name = match_fg_parameter_name(parameter_label)
    if not parameter_name or parameter_name not in allowed_parameters:
        return None

    parameter_name = canonicalize_fg_testing_method_name(parameter_name)
    if fieldname is None:
        fieldname = SINGLE_VALUE_FIELD_BY_PARAMETER.get(parameter_name, "target_value")

    return {
        "parameter": parameter_name,
        "fieldname": fieldname,
        "unit": resolve_parameter_unit(parameter_name, parameter_label),
    }


def get_limit_fieldname(raw_header: str) -> str | None:
    header = normalize_text(raw_header)
    if HEADER_MIN_PATTERN.search(header):
        return "minimum_value"
    if HEADER_MAX_PATTERN.search(header):
        return "maximum_value"
    return None


def strip_limit_tokens(raw_header: str) -> str:
    header = normalize_text(raw_header)
    header = HEADER_MIN_PATTERN.sub("", header)
    header = HEADER_MAX_PATTERN.sub("", header)
    header = re.sub(r"\(\s*\)", "", header)
    header = re.sub(r"\s+", " ", header).strip(" -")
    return header


def match_fg_parameter_name(parameter_label: str) -> str:
    cleaned_label = normalize_text(parameter_label)
    if not cleaned_label:
        return ""

    exact_key = cleaned_label.casefold()
    if exact_key in {"l", "a", "b"}:
        return exact_key.upper()

    canonical = canonicalize_fg_testing_method_name(cleaned_label)
    if canonical in get_true_fg_testing_method_name_set():
        return canonical

    for pattern, parameter_name in HEADER_PARAMETER_PATTERNS:
        if pattern.search(cleaned_label):
            return parameter_name
    return ""


def resolve_parameter_unit(parameter_name: str, parameter_label: str) -> str:
    unit = DEFAULT_PARAMETER_UNITS.get(parameter_name)
    if unit:
        return unit

    parenthetical_values = re.findall(r"\(([^()]*)\)", parameter_label)
    for candidate in parenthetical_values:
        cleaned = normalize_text(candidate)
        if not cleaned or cleaned.lower() in {"min", "max"}:
            continue
        return cleaned
    return ""


def build_item_standard_map(
    data_rows: list[tuple[object, ...]],
    parameter_columns: list[dict[str, object]],
) -> tuple[dict[str, dict[str, dict[str, object]]], list[str], list[str]]:
    item_standards: dict[str, dict[str, dict[str, object]]] = {}
    unmatched_items: list[str] = []
    source_item_counts: Counter[str] = Counter()

    for row in data_rows:
        values = list(row)
        if not any(normalize_text(value) for value in values):
            continue

        item_code = normalize_text(get_value(values, ITEM_CODE_INDEX)).upper()
        if not item_code:
            continue

        source_item_counts[item_code] += 1
        if not frappe.db.exists("Item", item_code):
            if item_code not in unmatched_items:
                unmatched_items.append(item_code)
            continue

        parameter_map = item_standards.setdefault(item_code, {})
        for column in parameter_columns:
            raw_value = get_value(values, column["index"])
            if raw_value in (None, ""):
                continue

            parameter_record = parameter_map.setdefault(
                column["parameter"],
                {
                    "minimum_value": None,
                    "maximum_value": None,
                    "target_value": "",
                    "unit": column["unit"],
                    "source_headers": set(),
                },
            )
            parameter_record["source_headers"].add(column["header"])
            if column["unit"] and not parameter_record.get("unit"):
                parameter_record["unit"] = column["unit"]

            assign_standard_value(parameter_record, column["fieldname"], raw_value)

    repeated_items = sorted(item_code for item_code, count in source_item_counts.items() if count > 1)
    return item_standards, unmatched_items, repeated_items


def assign_standard_value(parameter_record: dict[str, object], fieldname: str, raw_value) -> None:
    normalized_raw = normalize_text(raw_value)
    if not normalized_raw or normalized_raw.lower() in {"-", "n/a", "na", "nil"}:
        return

    if fieldname in {"minimum_value", "maximum_value"}:
        numeric_value = parse_float(normalized_raw)
        if numeric_value is not None:
            if abs(float(numeric_value)) > MAX_STORABLE_FG_LIMIT:
                if not parameter_record.get("target_value"):
                    parameter_record["target_value"] = normalized_raw
                return
            parameter_record[fieldname] = float(numeric_value)
            return
        parameter_record["target_value"] = normalized_raw
        return

    if not parameter_record.get("target_value"):
        parameter_record["target_value"] = normalized_raw


def sync_fg_parameter_standards(item_standards: dict[str, dict[str, dict[str, object]]]) -> dict[str, object]:
    item_codes = sorted(item_standards)
    if not item_codes:
        return {
            "grades_updated": 0,
            "rows_updated": 0,
            "rows_with_numeric_limits": 0,
            "rows_with_target_values": 0,
            "missing_standard_rows": [],
        }

    existing_rows = frappe.get_all(
        "FG Control Plan",
        filters={"fg_item_code": ["in", item_codes]},
        fields=[
            "name",
            "fg_item_code",
            "parameter",
            "applicable",
            "is_active",
            "test_type",
            "minimum_value",
            "maximum_value",
            "target_value",
            "unit",
        ],
        order_by="modified desc, creation desc",
        limit_page_length=0,
    )

    rows_updated = 0
    rows_with_numeric_limits = 0
    rows_with_target_values = 0
    grades_updated: set[str] = set()
    missing_standard_rows: list[str] = []

    for row in existing_rows:
        standards = item_standards.get(row["fg_item_code"], {})
        parameter_standard = standards.get(row["parameter"])
        doc = frappe.get_doc("FG Control Plan", row["name"])

        next_min = parameter_standard.get("minimum_value") if parameter_standard else None
        next_max = parameter_standard.get("maximum_value") if parameter_standard else None
        next_target = normalize_text(parameter_standard.get("target_value")) if parameter_standard else ""
        next_unit = normalize_text(parameter_standard.get("unit")) if parameter_standard else ""
        next_test_type = determine_fg_test_type(doc.parameter, next_min, next_max, next_target, doc.test_type)

        changed = False
        for fieldname, next_value in (
            ("minimum_value", next_min),
            ("maximum_value", next_max),
            ("target_value", next_target),
            ("unit", next_unit),
            ("test_type", next_test_type),
        ):
            current_value = doc.get(fieldname)
            if normalize_db_value(current_value) == normalize_db_value(next_value):
                continue
            doc.set(fieldname, next_value)
            changed = True

        if changed:
            doc.save(ignore_permissions=True)
            rows_updated += 1
            grades_updated.add(doc.fg_item_code)
            if rows_updated % FG_STANDARD_COMMIT_BATCH_SIZE == 0:
                frappe.db.commit()

        if doc.applicable and doc.is_active:
            if next_min is not None or next_max is not None:
                rows_with_numeric_limits += 1
            elif next_target:
                rows_with_target_values += 1
            else:
                missing_standard_rows.append(f"{doc.fg_item_code} :: {doc.parameter}")

    return {
        "grades_updated": len(grades_updated),
        "rows_updated": rows_updated,
        "rows_with_numeric_limits": rows_with_numeric_limits,
        "rows_with_target_values": rows_with_target_values,
        "missing_standard_rows": missing_standard_rows,
    }


def determine_fg_test_type(
    parameter: str,
    minimum_value,
    maximum_value,
    target_value: str,
    current_test_type: str,
) -> str:
    parameter_name = canonicalize_fg_testing_method_name(parameter)
    if parameter_name in FG_MANUAL_REVIEW_PARAMETERS:
        return "Manual"
    if minimum_value is not None or maximum_value is not None:
        return "Numeric"
    if target_value:
        return "Manual"
    return normalize_text(current_test_type) or "Manual"


def normalize_db_value(value):
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return normalize_text(value)


def get_value(values: list[object], index: int):
    if index >= len(values):
        return None
    return values[index]
