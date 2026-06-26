from __future__ import annotations

from collections import Counter
from pathlib import Path
from time import sleep

import frappe
from frappe.exceptions import QueryTimeoutError
from openpyxl import load_workbook

from calco_erp.calco_quality.fg_quality_setup import (
    FG_CONTROL_PLAN_IMPORT_DESCRIPTION,
    canonicalize_fg_testing_method_name,
    ensure_fg_quality_setup,
    get_true_fg_testing_method_names,
    get_true_fg_testing_method_name_set,
    has_positive_fg_control_requirement,
)
from calco_erp.calco_quality.rm_testing_utils import normalize_text, parse_float
from calco_erp.foundation_setup import QUALITY_PARAMETER_GROUP, ensure_quality_parameter_group


CONTROL_PLAN_NO_INDEX = 0
PRODUCT_CODE_INDEX = 1
CONTROL_PLAN_COMMIT_BATCH_SIZE = 500


def import_workbook(path: str) -> dict[str, object]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        frappe.throw(f"Workbook not found at {workbook_path}")

    ensure_quality_parameter_group()
    ensure_fg_quality_setup()

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return {"parameter_groups": 0, "items_processed": 0, "rows_created": 0, "rows_updated": 0}

    allowed_parameter_names = get_true_fg_testing_method_names()
    header = [normalize_text(value) for value in rows[0]]
    parameter_groups = derive_parameter_groups(header, allowed_parameter_names)
    parameter_names = unique_parameter_names(parameter_groups)

    master_summary = ensure_fg_parameter_master(parameter_names)
    sync_summary = sync_fg_control_plan_rows(rows[1:], parameter_groups)
    cleanup_summary = cleanup_non_quality_fg_artifacts(list(get_true_fg_testing_method_name_set()))
    frappe.clear_cache()
    frappe.db.commit()

    return {
        "parameter_groups": len(parameter_groups),
        "parameter_names": parameter_names,
        "fg_testing_methods": master_summary["fg_testing_methods"],
        "quality_parameters": master_summary["quality_parameters"],
        **sync_summary,
        **cleanup_summary,
    }


def derive_parameter_groups(header: list[str], allowed_parameter_names: tuple[str, ...] | list[str]) -> list[dict[str, object]]:
    groups: list[dict[str, object]] = []
    pending_parameters: list[dict[str, object]] = []
    allowed_names = set(allowed_parameter_names)
    index = PRODUCT_CODE_INDEX + 1

    while index < len(header):
        label = normalize_text(header[index])
        if not label:
            index += 1
            continue

        descriptive_size_base = get_descriptive_size_base(label)
        if descriptive_size_base:
            if pending_parameters and normalize_text(pending_parameters[-1]["name"]) == descriptive_size_base:
                frequency_index = None
                if index + 1 < len(header):
                    next_base = get_descriptive_frequency_base(header[index + 1])
                    if next_base and next_base == descriptive_size_base:
                        frequency_index = index + 1

                group = build_parameter_group(
                    pending_parameters,
                    index,
                    frequency_index,
                    allowed_names,
                )
                if group:
                    groups.append(group)
                pending_parameters.clear()
                index = (frequency_index + 1) if frequency_index is not None else (index + 1)
                continue

            index += 1
            continue

        if is_size_header(label):
            if pending_parameters:
                frequency_index = index + 1 if index + 1 < len(header) and is_frequency_header(header[index + 1]) else None
                group = build_parameter_group(
                    pending_parameters,
                    index,
                    frequency_index,
                    allowed_names,
                )
                if group:
                    groups.append(group)
                pending_parameters.clear()
                index = (frequency_index + 1) if frequency_index is not None else (index + 1)
                continue

            index += 1
            continue

        if is_frequency_header(label):
            index += 1
            continue

        if get_descriptive_frequency_base(label):
            index += 1
            continue

        pending_parameters.append({"name": canonicalize_fg_testing_method_name(label), "header_index": index})
        index += 1

    return groups


def build_parameter_group(
    pending_parameters: list[dict[str, object]],
    size_index: int,
    frequency_index: int | None,
    allowed_parameter_names: set[str],
) -> dict[str, object] | None:
    allowed_parameters = [parameter for parameter in pending_parameters if parameter["name"] in allowed_parameter_names]
    if not allowed_parameters:
        return None

    return {
        "parameters": allowed_parameters,
        "size_index": size_index,
        "frequency_index": frequency_index,
    }


def unique_parameter_names(parameter_groups: list[dict[str, object]]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for group in parameter_groups:
        for parameter in group["parameters"]:
            name = normalize_text(parameter["name"])
            if not name or name in seen:
                continue
            seen.add(name)
            names.append(name)
    return names


def ensure_fg_parameter_master(parameter_names: list[str]) -> dict[str, int]:
    fg_created = 0
    qi_created = 0
    existing_fg_methods = set()
    existing_quality_parameters = set()

    if frappe.db.exists("DocType", "FG Testing Method"):
        existing_fg_methods = set(frappe.get_all("FG Testing Method", pluck="name", limit_page_length=0))

    if frappe.db.exists("DocType", "Quality Inspection Parameter"):
        existing_quality_parameters = set(frappe.get_all("Quality Inspection Parameter", pluck="name", limit_page_length=0))

    for parameter_name in parameter_names:
        if parameter_name not in existing_fg_methods:
            insert_doc_with_retry(
                "FG Testing Method",
                parameter_name,
                {
                    "doctype": "FG Testing Method",
                    "method_name": parameter_name,
                },
            )
            existing_fg_methods.add(parameter_name)
            fg_created += 1

        if parameter_name not in existing_quality_parameters:
            doc = frappe.get_doc(
                {
                    "doctype": "Quality Inspection Parameter",
                    "parameter": parameter_name,
                    "parameter_group": QUALITY_PARAMETER_GROUP,
                    "description": FG_CONTROL_PLAN_IMPORT_DESCRIPTION,
                }
            )
            if doc.meta.has_field("custom_result_type"):
                doc.custom_result_type = "Manual Review"
            insert_doc_with_retry("Quality Inspection Parameter", parameter_name, doc)
            existing_quality_parameters.add(parameter_name)
            qi_created += 1

    return {
        "fg_testing_methods": fg_created,
        "quality_parameters": qi_created,
    }


def insert_doc_with_retry(doctype: str, record_name: str, payload, attempts: int = 3):
    for attempt in range(1, attempts + 1):
        if frappe.db.exists(doctype, record_name):
            return

        try:
            if hasattr(payload, "insert"):
                doc = frappe.get_doc(payload.as_dict())
            else:
                doc = frappe.get_doc(payload)
            doc.insert(ignore_permissions=True)
            frappe.db.commit()
            return
        except QueryTimeoutError:
            frappe.db.rollback()
            if attempt >= attempts:
                raise
            sleep(attempt)


def cleanup_non_quality_fg_artifacts(valid_parameter_names: list[str]) -> dict[str, int]:
    valid_names = set(valid_parameter_names)
    deleted_control_plan_rows = delete_invalid_fg_control_plan_rows(valid_names)
    deleted_template_rows = remove_invalid_template_rows(valid_names)
    deleted_fg_testing_methods = delete_invalid_fg_testing_methods(valid_names)
    deleted_quality_parameters = delete_invalid_quality_parameters(valid_names)

    return {
        "deleted_invalid_fg_control_plan_rows": deleted_control_plan_rows,
        "deleted_invalid_template_rows": deleted_template_rows,
        "deleted_invalid_fg_testing_methods": deleted_fg_testing_methods,
        "deleted_invalid_quality_parameters": deleted_quality_parameters,
    }


def delete_invalid_fg_control_plan_rows(valid_parameter_names: set[str]) -> int:
    invalid_rows = frappe.get_all(
        "FG Control Plan",
        filters={"parameter": ["not in", list(valid_parameter_names)]},
        pluck="name",
        limit_page_length=0,
    )
    if not invalid_rows:
        return 0

    for batch in chunked(invalid_rows, CONTROL_PLAN_COMMIT_BATCH_SIZE):
        frappe.db.delete("FG Control Plan", {"name": ["in", batch]})
        frappe.db.commit()

    return len(invalid_rows)


def remove_invalid_template_rows(valid_parameter_names: set[str]) -> int:
    imported_invalid_parameters = frappe.get_all(
        "Quality Inspection Parameter",
        filters={
            "description": FG_CONTROL_PLAN_IMPORT_DESCRIPTION,
            "name": ["not in", list(valid_parameter_names)],
        },
        pluck="name",
        limit_page_length=0,
    )
    if not imported_invalid_parameters:
        return 0

    rows_by_parent: dict[str, set[str]] = {}
    for row in frappe.get_all(
        "Item Quality Inspection Parameter",
        filters={"specification": ["in", imported_invalid_parameters]},
        fields=["name", "parent"],
        limit_page_length=0,
    ):
        rows_by_parent.setdefault(row["parent"], set()).add(row["name"])

    deleted_rows = 0
    for parent, row_names in rows_by_parent.items():
        template = frappe.get_doc("Quality Inspection Template", parent)
        remaining_rows = []
        for child in template.item_quality_inspection_parameter:
            if child.name in row_names:
                deleted_rows += 1
                continue
            remaining_rows.append(child)
        template.set("item_quality_inspection_parameter", remaining_rows)
        template.save(ignore_permissions=True)

    if deleted_rows:
        frappe.db.commit()

    return deleted_rows


def delete_invalid_fg_testing_methods(valid_parameter_names: set[str]) -> int:
    invalid_methods = frappe.get_all(
        "FG Testing Method",
        filters={"name": ["not in", list(valid_parameter_names)]},
        pluck="name",
        limit_page_length=0,
    )
    deleted = 0

    for method_name in invalid_methods:
        if frappe.db.exists("FG Control Plan", {"parameter": method_name}):
            continue
        frappe.delete_doc("FG Testing Method", method_name, ignore_permissions=True, force=1)
        deleted += 1

    if deleted:
        frappe.db.commit()

    return deleted


def delete_invalid_quality_parameters(valid_parameter_names: set[str]) -> int:
    invalid_parameters = frappe.get_all(
        "Quality Inspection Parameter",
        filters={
            "description": FG_CONTROL_PLAN_IMPORT_DESCRIPTION,
            "name": ["not in", list(valid_parameter_names)],
        },
        pluck="name",
        limit_page_length=0,
    )
    deleted = 0

    for parameter_name in invalid_parameters:
        if frappe.db.exists("Item Quality Inspection Parameter", {"specification": parameter_name}):
            continue
        if frappe.db.exists("Quality Inspection Reading", {"specification": parameter_name}):
            continue
        frappe.delete_doc("Quality Inspection Parameter", parameter_name, ignore_permissions=True, force=1)
        deleted += 1

    if deleted:
        frappe.db.commit()

    return deleted


def chunked(values: list[str], size: int):
    for index in range(0, len(values), size):
        yield values[index : index + size]


def sync_fg_control_plan_rows(
    data_rows: list[tuple[object, ...]],
    parameter_groups: list[dict[str, object]],
) -> dict[str, object]:
    imported_rows_by_key: dict[tuple[str, str], dict[str, object]] = {}
    imported_item_codes: set[str] = set()
    unmatched_items: list[str] = []
    source_item_counts: Counter[str] = Counter()

    for row in data_rows:
        values = list(row)
        if not any(normalize_text(value) for value in values):
            continue

        item_code = normalize_text(get_value(values, PRODUCT_CODE_INDEX))
        control_plan_no = normalize_text(get_value(values, CONTROL_PLAN_NO_INDEX))
        if not item_code:
            continue

        source_item_counts[item_code] += 1
        if not frappe.db.exists("Item", item_code):
            if item_code not in unmatched_items:
                unmatched_items.append(item_code)
            continue

        imported_item_codes.add(item_code)
        for group in parameter_groups:
            size_value = parse_size_value(get_value(values, group["size_index"]))
            frequency_value = normalize_text(get_value(values, group["frequency_index"])) if group["frequency_index"] is not None else ""
            applicable = 1 if has_positive_fg_control_requirement(size_value, frequency_value) else 0

            for parameter in group["parameters"]:
                imported_rows_by_key[(item_code, parameter["name"])] = {
                    "fg_item_code": item_code,
                    "parameter": parameter["name"],
                    "applicable": applicable,
                    "size": size_value,
                    "frequency": frequency_value,
                    "version": control_plan_no or "1.0",
                }

    existing_rows = {}
    existing_duplicates = 0
    for row in frappe.get_all(
        "FG Control Plan",
        filters={"fg_item_code": ["in", list(imported_item_codes or [""])]},
        fields=[
            "name",
            "fg_item_code",
            "parameter",
            "critical_test",
            "test_type",
            "minimum_value",
            "maximum_value",
            "target_value",
            "unit",
            "modified",
        ],
        order_by="modified desc, creation desc",
        limit_page_length=0,
    ):
        key = (row["fg_item_code"], row["parameter"])
        if key in existing_rows:
            existing_duplicates += 1
            frappe.db.set_value("FG Control Plan", row["name"], "is_active", 0, update_modified=False)
            frappe.db.set_value("FG Control Plan", row["name"], "applicable", 0, update_modified=False)
            continue
        existing_rows[key] = row

    created = 0
    updated = 0
    imported_keys = set(imported_rows_by_key)
    write_count = 0

    for row in imported_rows_by_key.values():
        key = (row["fg_item_code"], row["parameter"])
        existing = existing_rows.get(key)
        if not existing:
            existing_name = frappe.db.get_value(
                "FG Control Plan",
                {
                    "fg_item_code": row["fg_item_code"],
                    "parameter": row["parameter"],
                },
                "name",
            )
            if existing_name:
                existing = {
                    "name": existing_name,
                    "fg_item_code": row["fg_item_code"],
                    "parameter": row["parameter"],
                    "critical_test": frappe.db.get_value("FG Control Plan", existing_name, "critical_test"),
                    "test_type": frappe.db.get_value("FG Control Plan", existing_name, "test_type"),
                    "minimum_value": frappe.db.get_value("FG Control Plan", existing_name, "minimum_value"),
                    "maximum_value": frappe.db.get_value("FG Control Plan", existing_name, "maximum_value"),
                    "target_value": frappe.db.get_value("FG Control Plan", existing_name, "target_value"),
                    "unit": frappe.db.get_value("FG Control Plan", existing_name, "unit"),
                }
                existing_rows[key] = existing

        if existing:
            doc = frappe.get_doc("FG Control Plan", existing["name"])
            updated += 1
        else:
            doc = frappe.new_doc("FG Control Plan")
            doc.parameter = row["parameter"]
            doc.test_type = "Manual"
            doc.critical_test = 0
            created += 1

        doc.fg_item_code = row["fg_item_code"]
        doc.parameter = row["parameter"]
        doc.applicable = row["applicable"]
        doc.size = row["size"]
        doc.frequency = row["frequency"]
        doc.version = row["version"]
        doc.is_active = 1

        if existing:
            doc.critical_test = existing["critical_test"]
            doc.test_type = normalize_text(existing["test_type"]) or "Manual"
            doc.minimum_value = existing["minimum_value"]
            doc.maximum_value = existing["maximum_value"]
            doc.target_value = existing.get("target_value")
            doc.unit = existing.get("unit")

        if doc.is_new():
            doc.insert(ignore_permissions=True)
        else:
            doc.save(ignore_permissions=True)
        write_count += 1
        if write_count % CONTROL_PLAN_COMMIT_BATCH_SIZE == 0:
            frappe.db.commit()

    deactivated = 0
    for key, row in existing_rows.items():
        if key in imported_keys:
            continue
        frappe.db.set_value("FG Control Plan", row["name"], "is_active", 0, update_modified=False)
        frappe.db.set_value("FG Control Plan", row["name"], "applicable", 0, update_modified=False)
        deactivated += 1
        write_count += 1
        if write_count % CONTROL_PLAN_COMMIT_BATCH_SIZE == 0:
            frappe.db.commit()

    applicable_count = sum(1 for row in imported_rows_by_key.values() if row["applicable"])
    repeated_source_items = sorted(item_code for item_code, count in source_item_counts.items() if count > 1)

    return {
        "items_processed": len(imported_item_codes),
        "rows_created": created,
        "rows_updated": updated,
        "rows_deactivated": deactivated,
        "existing_duplicates_deactivated": existing_duplicates,
        "applicable_rows": applicable_count,
        "repeated_source_items": repeated_source_items,
        "unmatched_items": unmatched_items,
    }


def is_size_header(value: str) -> bool:
    return normalize_text(value).lower().startswith("size")


def is_frequency_header(value: str) -> bool:
    return normalize_text(value).lower().startswith("frequency")


def get_descriptive_size_base(value: str) -> str:
    label = normalize_text(value)
    suffix = " - Size"
    if label.endswith(suffix):
        return canonicalize_fg_testing_method_name(label[: -len(suffix)])
    return ""


def get_descriptive_frequency_base(value: str) -> str:
    label = normalize_text(value)
    suffix = " - Frequency"
    if label.endswith(suffix):
        return canonicalize_fg_testing_method_name(label[: -len(suffix)])
    return ""


def parse_size_value(value) -> float | None:
    normalized = normalize_text(value)
    if not normalized or normalized.lower() in {"n/a", "na"}:
        return None

    parsed = parse_float(normalized)
    if parsed is None:
        return None
    return float(parsed)


def get_value(values: list[object], index: int | None):
    if index is None or index >= len(values):
        return None
    return values[index]
